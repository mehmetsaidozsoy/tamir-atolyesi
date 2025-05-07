import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from database import VeritabaniYonetici
from ttkthemes import ThemedStyle
from datetime import datetime, timedelta, timezone
import json
import os
import matplotlib
matplotlib.use('Agg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import math
import yagmail
import time
import tzlocal
from zoneinfo import ZoneInfo
from PIL import Image, ImageTk

logger = logging.getLogger(__name__)

class TamirAtolyesiGUI:
    # Cihaz kategorileri
    cihaz_kategorileri = {
        "Bilgisayar": ["Masaüstü", "Dizüstü", "All-in-One"],
        "Telefon": ["Akıllı Telefon", "Tablet"],
        "Yazıcı": ["Lazer Yazıcı", "Mürekkep Püskürtmeli", "Nokta Vuruşlu"],
        "Televizyon": ["LCD", "LED", "OLED", "Plazma"],
        "Beyaz Eşya": ["Buzdolabı", "Çamaşır Makinesi", "Bulaşık Makinesi", "Fırın"],
        "Küçük Ev Aletleri": ["Mikser", "Blender", "Tost Makinesi", "Ütü"],
        "Diğer": ["Diğer"]
    }
    
    def __init__(self, parent, current_user=None, current_user_yetki_seviyesi=None):
        """Ana uygulamayı başlatır"""
        self.parent = parent
        self.db = VeritabaniYonetici()
        self.current_user = current_user if current_user else 'admin'
        self.current_user_yetki_seviyesi = int(current_user_yetki_seviyesi) if current_user_yetki_seviyesi is not None else 0
        
        # Notebook (sekmeler) oluştur
        self.notebook = ttk.Notebook(self.parent)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Sekmeleri oluştur
        self.setup_dashboard_tab()
        self.setup_repairs_tab()
        self.setup_customers_tab()
        self.setup_reports_tab()
        self.setup_settings_tab()
        self.setup_stats_tab()
        if self.current_user_yetki_seviyesi == 0:
            self.setup_admin_panel_tab()  # Yönetim Paneli sadece admin için
        self.setup_about_tab()  # Hakkında sekmesi
        self.setup_manual_tab() # Kullanım Bilgisi sekmesi
        
        # İstatistikleri güncelle
        self.update_dashboard()
        
        # Otomatik yedekleme zamanlayıcısını başlat
        self.check_auto_backup()
        
    def setup_main_window(self):
        """Ana pencereyi hazırlar"""
        # Ana frame
        self.main_frame = ttk.Frame(self.parent, padding="10")
        self.main_frame.pack(fill="both", expand=True)
        
        # Menü çubuğu
        self.create_menu()
        
        # Araç çubuğu
        self.create_toolbar()
        
        # Notebook (sekmeler)
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill="both", expand=True, pady=(10, 0))
        
        # Sekmeleri oluştur
        self.create_tabs()
        
        # Durum çubuğu
        self.statusbar = ttk.Label(self.parent, text="Hazır", relief="sunken")
        self.statusbar.pack(side="bottom", fill="x")
        
    def create_menu(self):
        """Menü çubuğunu oluşturur"""
        menubar = tk.Menu(self.parent)
        self.parent.config(menu=menubar)
        
        # Dosya menüsü
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Dosya", menu=file_menu)
        file_menu.add_command(label="Yeni Müşteri", command=self.new_customer)
        file_menu.add_command(label="Yeni Tamir", command=self.new_repair)
        file_menu.add_separator()
        file_menu.add_command(label="Çıkış", command=self.parent.quit)
        
        # Araçlar menüsü
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Araçlar", menu=tools_menu)
        tools_menu.add_command(label="Yedekleme Yöneticisi", command=self.show_backup_manager)
        
        # Raporlar menüsü
        reports_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Raporlar", menu=reports_menu)
        reports_menu.add_command(label="Müşteri Raporu", command=self.show_customer_report)
        reports_menu.add_command(label="Tamir Raporu", command=self.show_repair_report)
        reports_menu.add_command(label="Gelir Raporu", command=self.show_income_report)
        
        # İstatistikler menüsü
        stats_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="İstatistikler", menu=stats_menu)
        stats_menu.add_command(label="Aylık İstatistikler", command=self.show_monthly_stats)
        stats_menu.add_command(label="Yıllık İstatistikler", command=self.show_yearly_stats)
        
        # Yardım menüsü
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Yardım", menu=help_menu)
        help_menu.add_command(label="Kullanıcı Kılavuzu", command=self.show_user_manual)
        help_menu.add_command(label="Hakkında", command=self.show_about)
        
    def create_toolbar(self):
        """Araç çubuğunu oluşturur"""
        toolbar = ttk.Frame(self.main_frame)
        toolbar.pack(fill="x", pady=(0, 10))
        
        ttk.Button(toolbar, text="Yeni Müşteri", command=self.new_customer).pack(side="left", padx=2)
        ttk.Button(toolbar, text="Yeni Tamir", command=self.new_repair).pack(side="left", padx=2)
        ttk.Button(toolbar, text="Yedekleme Yöneticisi", command=self.show_backup_manager).pack(side="left", padx=2)
        
    def create_tabs(self):
        """Sekmeleri oluşturur"""
        # Müşteriler sekmesi
        self.customers_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.customers_frame, text="Müşteriler")
        self.setup_customers_tab()
        
        # Tamirler sekmesi
        self.repairs_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.repairs_frame, text="Tamirler")
        self.setup_repairs_tab()
        
        # Ayarlar sekmesi
        self.settings_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.settings_frame, text="Ayarlar")
        self.setup_settings_tab()
        
        # Ana sayfa sekmesi
        self.dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.dashboard_frame, text="Ana Sayfa")
        self.setup_dashboard_tab()
        
        # Raporlar sekmesi
        self.reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.reports_frame, text="Raporlar")
        self.setup_reports_tab()
        
    def setup_customers_tab(self):
        """Müşteriler sekmesini hazırlar"""
        # Müşteriler sekmesi frame'i
        self.customers_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.customers_frame, text="Müşteriler")
        
        # Üst kısım - Arama
        top_frame = ttk.Frame(self.customers_frame)
        top_frame.pack(fill="x", pady=(0, 10))
        
        # Arama çubuğu
        search_frame = ttk.LabelFrame(top_frame, text="Arama", padding="5")
        search_frame.pack(side="left", fill="x", expand=True)
        
        ttk.Label(search_frame, text="Arama:").pack(side="left", padx=5)
        self.customer_search = ttk.Entry(search_frame)
        self.customer_search.pack(side="left", fill="x", expand=True, padx=5)
        
        ttk.Button(
            search_frame,
            text="Ara",
            command=self.search_customers
        ).pack(side="left", padx=5)
        
        # Yeni müşteri butonu
        ttk.Button(
            top_frame,
            text="Yeni Müşteri",
            command=lambda: self.new_customer()
        ).pack(side="right", padx=5)
        
        # Müşteri listesi
        self.customers_tree = ttk.Treeview(
            self.customers_frame,
            columns=("id", "ad_soyad", "telefon", "email", "adres"),
            show="headings"
        )
        
        # Sütun başlıkları
        self.customers_tree.heading("id", text="ID")
        self.customers_tree.heading("ad_soyad", text="Ad Soyad")
        self.customers_tree.heading("telefon", text="Telefon")
        self.customers_tree.heading("email", text="E-posta")
        self.customers_tree.heading("adres", text="Adres")
        
        # Sütun genişlikleri
        self.customers_tree.column("id", width=50)
        self.customers_tree.column("ad_soyad", width=200)
        self.customers_tree.column("telefon", width=100)
        self.customers_tree.column("email", width=150)
        self.customers_tree.column("adres", width=300)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.customers_frame, orient="vertical", command=self.customers_tree.yview)
        self.customers_tree.configure(yscrollcommand=scrollbar.set)
        
        # Yerleştirme
        self.customers_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Sağ tık menüsü
        self.customer_menu = tk.Menu(self.parent, tearoff=0)
        self.customer_menu.add_command(label="Düzenle", command=self.edit_customer)
        self.customer_menu.add_command(label="Sil", command=self.delete_customer)
        
        # Sağ tık olayı
        self.customers_tree.bind("<Button-3>", self.show_customer_menu)
        
        # Listeyi güncelle
        self.refresh_customers()

    def search_customers(self):
        """Müşteri arama"""
        try:
            arama = self.customer_search.get().strip()
            if not arama:
                self.refresh_customers()
                return
                
            # Mevcut listeyi temizle
            for item in self.customers_tree.get_children():
                self.customers_tree.delete(item)
                
            # Müşterileri ara ve listele
            musteriler = self.db.musteri_ara(arama)
            for musteri in musteriler:
                self.customers_tree.insert("", "end", values=(
                    musteri[0],  # ID
                    f"{musteri[2]} {musteri[3]}",  # Ad Soyad
                    musteri[4],  # Telefon
                    musteri[5],  # Email
                    musteri[6]   # Adres
                ))
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def refresh_customers(self):
        """Müşteri listesini günceller"""
        try:
            # Mevcut listeyi temizle
            for item in self.customers_tree.get_children():
                self.customers_tree.delete(item)
                
            # Müşterileri getir ve listele
            musteriler = self.db.musteri_listesi()
            for musteri in musteriler:
                self.customers_tree.insert("", "end", values=(
                    musteri['id'],
                    f"{musteri['ad']} {musteri['soyad']}",
                    musteri['telefon'],
                    musteri['eposta'],
                    musteri['adres']
                ))
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def show_customer_menu(self, event):
        """Müşteri sağ tık menüsünü gösterir"""
        item = self.customers_tree.identify_row(event.y)
        if item:
            self.customers_tree.selection_set(item)
            self.customer_menu.post(event.x_root, event.y_root)

    def edit_customer(self):
        """Seçili müşteriyi düzenler"""
        item = self.customers_tree.selection()
        if not item:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek müşteriyi seçin!")
            return
        musteri_id = self.customers_tree.item(item[0])['values'][0]
        musteri = self.db.musteri_getir(musteri_id)
        if musteri:
            self.new_customer(edit_mode=True, musteri=musteri)
        else:
            messagebox.showerror("Hata", "Seçili müşteri veritabanında bulunamadı!")

    def delete_customer(self):
        """Seçili müşteriyi siler"""
        item = self.customers_tree.selection()
        if not item:
            return
        musteri_id = self.customers_tree.item(item[0])['values'][0]
        if messagebox.askyesno("Onay", "Bu müşteriyi silmek istediğinize emin misiniz?\nMüşteriye ait tüm tamir kayıtları da silinecektir."):
            try:
                self.db.musteri_sil(musteri_id)
                self.refresh_customers()
                messagebox.showinfo("Başarılı", "Müşteri silindi")
            except Exception as e:
                messagebox.showerror("Hata", str(e))

    def setup_repairs_tab(self):
        """Tamirler sekmesini hazırlar"""
        self.repairs_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.repairs_frame, text="Tamirler")
        
        # Üst kısım - Arama ve filtreler
        top_frame = ttk.Frame(self.repairs_frame)
        top_frame.pack(fill="x", pady=(0, 10))
        
        # Arama çubuğu
        search_frame = ttk.LabelFrame(top_frame, text="Arama", padding="5")
        search_frame.pack(side="left", fill="x", expand=True)
        
        ttk.Label(search_frame, text="Arama:").pack(side="left", padx=5)
        self.repair_search = ttk.Entry(search_frame)
        self.repair_search.pack(side="left", fill="x", expand=True, padx=5)
        
        ttk.Button(
            search_frame,
            text="Ara",
            command=self.search_repairs
        ).pack(side="left", padx=5)
        
        # Durum filtresi
        filter_frame = ttk.LabelFrame(top_frame, text="Filtrele", padding="5")
        filter_frame.pack(side="left", padx=10)
        
        ttk.Label(filter_frame, text="Durum:").pack(side="left", padx=5)
        self.repair_filter = ttk.Combobox(filter_frame, values=[
            "Tümü", "Beklemede", "Parça Bekleniyor", "Tamir Ediliyor",
            "Test Ediliyor", "Tamamlandı", "İptal Edildi"
        ])
        self.repair_filter.set("Tümü")
        self.repair_filter.pack(side="left", padx=5)
        self.repair_filter.bind('<<ComboboxSelected>>', lambda e: self.refresh_repairs())
        
        # Yeni tamir butonu
        ttk.Button(
            top_frame,
            text="Yeni Tamir",
            command=self.new_repair
        ).pack(side="right", padx=5)
        
        # Tamir listesi
        self.repairs_tree = ttk.Treeview(
            self.repairs_frame,
            columns=("id", "tarih", "musteri", "cihaz", "durum", "ucret"),
            show="headings"
        )
        
        # Sütun başlıkları
        self.repairs_tree.heading("id", text="ID")
        self.repairs_tree.heading("tarih", text="Tarih")
        self.repairs_tree.heading("musteri", text="Müşteri")
        self.repairs_tree.heading("cihaz", text="Cihaz")
        self.repairs_tree.heading("durum", text="Durum")
        self.repairs_tree.heading("ucret", text="Ücret")
        
        # Sütun genişlikleri
        self.repairs_tree.column("id", width=50)
        self.repairs_tree.column("tarih", width=100)
        self.repairs_tree.column("musteri", width=200)
        self.repairs_tree.column("cihaz", width=200)
        self.repairs_tree.column("durum", width=100)
        self.repairs_tree.column("ucret", width=100)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.repairs_frame, orient="vertical", command=self.repairs_tree.yview)
        self.repairs_tree.configure(yscrollcommand=scrollbar.set)
        
        # Yerleştirme
        self.repairs_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Sağ tık menüsü
        self.repair_menu = tk.Menu(self.parent, tearoff=0)
        self.repair_menu.add_command(label="Düzenle", command=self.edit_repair)
        self.repair_menu.add_command(label="Sil", command=self.delete_repair)
        self.repair_menu.add_separator()
        self.repair_menu.add_command(label="Durumu Güncelle", command=self.update_repair_status)
        
        # Sağ tık olayı
        self.repairs_tree.bind("<Button-3>", self.show_repair_menu)
        
        # Listeyi güncelle
        self.refresh_repairs()

    def search_repairs(self):
        """Tamir arama"""
        try:
            arama = self.repair_search.get().strip()
            if not arama:
                self.refresh_repairs()
                return
                
            # Mevcut listeyi temizle
            for item in self.repairs_tree.get_children():
                self.repairs_tree.delete(item)
                
            # Tamirleri ara ve listele
            tamirler = self.db.tamir_ara(arama)
            for tamir in tamirler:
                try:
                    # Toplam ücreti doğru sütundan al (tamir[6])
                    ucret = float(tamir[6]) if tamir[6] and str(tamir[6]).strip() != '' else 0
                    ucret_str = "{:.2f} TL".format(ucret)
                except (ValueError, TypeError):
                    ucret_str = "0.00 TL"
                    
                self.repairs_tree.insert("", "end", values=(
                    tamir[0],  # ID
                    tamir[1],  # Tarih (giris_tarihi veya tarih)
                    f"{tamir[2]} {tamir[3]}",  # Müşteri Adı Soyadı
                    tamir[4],  # Cihaz
                    tamir[5],  # Durum
                    ucret_str  # Toplam Ücret
                ))
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def refresh_repairs(self):
        """Tamir listesini günceller"""
        try:
            # Mevcut listeyi temizle
            for item in self.repairs_tree.get_children():
                self.repairs_tree.delete(item)
            # Tamirleri getir ve listele
            tamirler = self.db.tum_tamirler()
            for tamir in tamirler:
                try:
                    # Toplam ücreti doğru sütundan al (tamir[6])
                    ucret = float(tamir[6]) if tamir[6] and str(tamir[6]).strip() != '' else 0
                    ucret_str = "{:.2f} TL".format(ucret)
                except (ValueError, TypeError):
                    ucret_str = "0.00 TL"
                self.repairs_tree.insert("", "end", values=(
                    tamir[0],  # ID
                    tamir[1],  # Tarih (giris_tarihi veya tarih)
                    f"{tamir[2]} {tamir[3]}",  # Müşteri Adı Soyadı
                    tamir[4],  # Cihaz
                    tamir[5],  # Durum
                    ucret_str  # Toplam Ücret
                ))
        except Exception as e:
            messagebox.showerror("Hata", str(e))
        
    def show_repair_menu(self, event):
        """Tamir sağ tık menüsünü gösterir"""
        item = self.repairs_tree.selection()
        if item:
            self.repair_menu.post(event.x_root, event.y_root)
            
    def show_repair_details(self):
        """Seçili tamirin detaylarını gösterir"""
        selected = self.repairs_tree.selection()
        if not selected:
            return
            
        repair_id = self.repairs_tree.item(selected[0])['values'][0]
        repair = self.db.onarim_getir(repair_id)
        if not repair:
            return
            
        dialog = tk.Toplevel(self.parent)
        dialog.title("Tamir Detayları")
        dialog.geometry("600x800")
        dialog.resizable(False, False)
        
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Müşteri bilgileri
        customer_frame = ttk.LabelFrame(main_frame, text="Müşteri Bilgileri", padding="10")
        customer_frame.pack(fill="x", pady=(0, 10))
        
        musteri = self.db.musteri_getir(repair['musteri_id'])
        if musteri:
            ttk.Label(customer_frame, text=f"Ad Soyad: {musteri[2]} {musteri[3]}").pack(anchor="w")
            ttk.Label(customer_frame, text=f"Telefon: {musteri[4]}").pack(anchor="w")
            ttk.Label(customer_frame, text=f"E-posta: {musteri[5]}").pack(anchor="w")
        
        # Cihaz bilgileri
        device_frame = ttk.LabelFrame(main_frame, text="Cihaz Bilgileri", padding="10")
        device_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(device_frame, text=f"Cihaz Türü: {repair['cihaz_turu']}").pack(anchor="w")
        ttk.Label(device_frame, text=f"Marka: {repair['marka']}").pack(anchor="w")
        ttk.Label(device_frame, text=f"Model: {repair['model']}").pack(anchor="w")
        ttk.Label(device_frame, text=f"Seri No: {repair['seri_no']}").pack(anchor="w")
        
        # Arıza ve işlem bilgileri
        repair_frame = ttk.LabelFrame(main_frame, text="Arıza ve İşlem Bilgileri", padding="10")
        repair_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(repair_frame, text="Müşteri Şikayeti:").pack(anchor="w", pady=(0, 5))
        sikayet = tk.Text(repair_frame, height=4, width=50)
        sikayet.insert("1.0", repair['sikayet'])
        sikayet.configure(state="disabled")
        sikayet.pack(fill="x", pady=(0, 10))
        
        ttk.Label(repair_frame, text="Tespit Edilen Arıza:").pack(anchor="w", pady=(0, 5))
        ariza = tk.Text(repair_frame, height=4, width=50)
        ariza.insert("1.0", repair['ariza'])
        ariza.configure(state="disabled")
        ariza.pack(fill="x", pady=(0, 10))
        
        ttk.Label(repair_frame, text="Yapılan İşlemler:").pack(anchor="w", pady=(0, 5))
        islemler = tk.Text(repair_frame, height=4, width=50)
        islemler.insert("1.0", repair['islemler'])
        islemler.configure(state="disabled")
        islemler.pack(fill="x", pady=(0, 10))
        
        ttk.Label(repair_frame, text="Kullanılan Parçalar:").pack(anchor="w", pady=(0, 5))
        parcalar = tk.Text(repair_frame, height=4, width=50)
        parcalar.insert("1.0", repair['parcalar'])
        parcalar.configure(state="disabled")
        parcalar.pack(fill="x", pady=(0, 10))
        
        # Ücret bilgileri
        cost_frame = ttk.LabelFrame(main_frame, text="Ücret Bilgileri", padding="10")
        cost_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(cost_frame, text=f"İşçilik Ücreti: {repair['iscilik_ucreti']:.2f} TL").pack(anchor="w")
        ttk.Label(cost_frame, text=f"Parça Ücreti: {repair['parca_ucreti']:.2f} TL").pack(anchor="w")
        ttk.Label(cost_frame, text=f"Toplam Ücret: {repair['toplam_ucret']:.2f} TL").pack(anchor="w")
        
        # Durum bilgisi
        status_frame = ttk.LabelFrame(main_frame, text="Durum Bilgisi", padding="10")
        status_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(status_frame, text=f"Durum: {repair['durum']}").pack(anchor="w")
        ttk.Label(status_frame, text=f"Son Güncelleme: {repair['guncelleme_tarihi']}").pack(anchor="w")
        
        # Kapat butonu
        ttk.Button(main_frame, text="Kapat", command=dialog.destroy).pack(pady=10)
        
    def update_repair_status(self):
        """Seçili tamirin durumunu günceller"""
        item = self.repairs_tree.selection()
        if not item:
            return
        tamir_id = self.repairs_tree.item(item[0])['values'][0]
        tamir = self.db.tamir_getir(tamir_id)
        if not tamir:
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("Durum Güncelle")
        dialog.geometry("300x150")
        dialog.resizable(False, False)
        
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="Yeni Durum:").pack(pady=5)
        durum = ttk.Combobox(main_frame, values=[
            "Beklemede", "Parça Bekleniyor", "Tamir Ediliyor",
            "Test Ediliyor", "Tamamlandı", "İptal Edildi"
        ])
        durum.set(tamir['durum'])
        durum.pack(pady=5)
        
        def kaydet():
            try:
                self.db.tamir_durum_guncelle(tamir_id, durum.get())
                self.refresh_repairs()
                dialog.destroy()
                messagebox.showinfo("Başarılı", "Tamir durumu güncellendi")
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        
        ttk.Button(
            main_frame,
            text="Kaydet",
            command=kaydet
        ).pack(pady=10)

    def edit_repair(self):
        """Seçili tamiri düzenler"""
        item = self.repairs_tree.selection()
        if not item:
            return
        tamir_id = self.repairs_tree.item(item[0])['values'][0]
        tamir = self.db.tamir_getir(tamir_id)
        if tamir:
            self.new_repair(edit_mode=True, tamir=tamir)

    def delete_repair(self):
        """Seçili tamiri siler"""
        item = self.repairs_tree.selection()
        if not item:
            return
        tamir_id = self.repairs_tree.item(item[0])['values'][0]
        if messagebox.askyesno("Onay", "Bu tamir kaydını silmek istediğinize emin misiniz?"):
            try:
                self.db.tamir_sil(tamir_id)
                self.refresh_repairs()
                messagebox.showinfo("Başarılı", "Tamir kaydı silindi")
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        
    def setup_settings_tab(self):
        """Ayarlar sekmesini hazırlar"""
        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="Ayarlar")
        
        # Genel ayarlar
        general_frame = ttk.LabelFrame(self.settings_frame, text="Genel Ayarlar", padding="10")
        general_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(general_frame, text="Firma Adı:").grid(row=0, column=0, padx=5, pady=5)
        self.company_name = ttk.Entry(general_frame)
        self.company_name.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(general_frame, text="Adres:").grid(row=1, column=0, padx=5, pady=5)
        self.company_address = ttk.Entry(general_frame)
        self.company_address.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(general_frame, text="Telefon:").grid(row=2, column=0, padx=5, pady=5)
        self.company_phone = ttk.Entry(general_frame)
        self.company_phone.grid(row=2, column=1, padx=5, pady=5)
        
        # Yedekleme ayarları
        backup_frame = ttk.LabelFrame(self.settings_frame, text="Yedekleme", padding="10")
        backup_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(backup_frame, text="Otomatik Yedekleme:").grid(row=0, column=0, padx=5, pady=5)
        self.auto_backup = tk.BooleanVar()
        ttk.Checkbutton(
            backup_frame,
            text="Aktif",
            variable=self.auto_backup
        ).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(backup_frame, text="Yedekleme Aralığı (Gün):").grid(row=1, column=0, padx=5, pady=5)
        self.backup_interval = ttk.Entry(backup_frame)
        self.backup_interval.grid(row=1, column=1, padx=5, pady=5)
        self.backup_interval.insert(0, "7")
        
        button_frame = ttk.Frame(backup_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(
            button_frame,
            text="Yedek Al",
            command=self.create_backup
        ).pack(side="left", padx=5)
        
        ttk.Button(
            button_frame,
            text="Yedek Geri Yükle",
            command=self.restore_backup
        ).pack(side="left", padx=5)
        
        # Yedekleme Yöneticisi butonu
        ttk.Button(
            button_frame,
            text="Yedekleme Yöneticisi",
            command=self.show_backup_manager
        ).pack(side="left", padx=5)
        
        # Yedek listesi
        backup_list_frame = ttk.LabelFrame(self.settings_frame, text="Yedek Listesi", padding="10")
        backup_list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.backup_tree = ttk.Treeview(
            backup_list_frame,
            columns=("id", "tarih", "dosya_adi", "boyut", "aciklama"),
            show="headings"
        )
        
        # Sütun başlıkları
        self.backup_tree.heading("id", text="ID")
        self.backup_tree.heading("tarih", text="Tarih")
        self.backup_tree.heading("dosya_adi", text="Dosya Adı")
        self.backup_tree.heading("boyut", text="Boyut")
        self.backup_tree.heading("aciklama", text="Açıklama")
        
        # Sütun genişlikleri
        self.backup_tree.column("id", width=0, stretch=False)
        self.backup_tree.column("tarih", width=150)
        self.backup_tree.column("dosya_adi", width=220)
        self.backup_tree.column("boyut", width=100)
        self.backup_tree.column("aciklama", width=300)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(backup_list_frame, orient="vertical", command=self.backup_tree.yview)
        self.backup_tree.configure(yscrollcommand=scrollbar.set)
        
        # Yerleştirme
        self.backup_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Yedek listesini güncelle
        self.refresh_backups()
        
        # Mevcut ayarları yükle
        self.load_settings()
        
        # E-posta ayarları
        email_frame = ttk.LabelFrame(self.settings_frame, text="Yedek E-posta Ayarları", padding="10")
        email_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(email_frame, text="Gönderici E-posta:").grid(row=0, column=0, padx=5, pady=5)
        self.email_sender = ttk.Entry(email_frame, width=35)
        self.email_sender.grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(email_frame, text="Şifre:").grid(row=1, column=0, padx=5, pady=5)
        self.email_password = ttk.Entry(email_frame, show="*", width=35)
        self.email_password.grid(row=1, column=1, padx=5, pady=5)
        ttk.Label(email_frame, text="Alıcı E-posta:").grid(row=2, column=0, padx=5, pady=5)
        self.email_receiver = ttk.Entry(email_frame, width=35)
        self.email_receiver.grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(email_frame, text="Kaydet", command=self.save_email_settings).grid(row=3, column=0, columnspan=2, pady=5)
        # Ayarları yükle
        self.load_email_settings()
        # Test verisi ekle butonu
        # ttk.Button(self.settings_frame, text="Test Verisi Ekle", command=self.add_test_data).pack(pady=10)
        # giris_tarihi alanı ekle butonu
        # ttk.Button(self.settings_frame, text="giris_tarihi Alanı Ekle", command=self.ekle_giris_tarihi_alani).pack(pady=2)
        # Test verilerini sil butonu
        # ttk.Button(self.settings_frame, text="Test Verilerini Sil", command=self.sil_test_tamirleri).pack(pady=2)
        # Eksik giris_tarihi doldur butonu
        # ttk.Button(self.settings_frame, text="Eksik giris_tarihi Doldur", command=self.doldur_eksik_giris_tarihi).pack(pady=2)
        # ttk.Button(self.settings_frame, text="Tüm giris_tarihi 3 saat ileri al", command=self.giris_tarihi_ileri_al).pack(pady=2)

    def load_settings(self):
        """Mevcut ayarları yükler"""
        try:
            # Firma bilgileri
            firma_adi = self.db.ayar_getir("firma_adi")
            if firma_adi:
                self.company_name.insert(0, firma_adi)
                
            firma_adres = self.db.ayar_getir("firma_adres")
            if firma_adres:
                self.company_address.insert(0, firma_adres)
                
            firma_telefon = self.db.ayar_getir("firma_telefon")
            if firma_telefon:
                self.company_phone.insert(0, firma_telefon)
                
            # Yedekleme ayarları
            auto_backup = self.db.ayar_getir("auto_backup")
            if auto_backup:
                self.auto_backup.set(auto_backup.lower() == "true")
                
            backup_interval = self.db.ayar_getir("backup_interval")
            if backup_interval:
                self.backup_interval.delete(0, "end")
                self.backup_interval.insert(0, backup_interval)
                
        except Exception as e:
            logger.error(f"Ayarlar yüklenirken hata: {str(e)}")

    def save_settings(self):
        """Ayarları kaydeder"""
        try:
            # Ayarları veritabanına kaydet
            self.db.ayar_ekle("auto_backup", str(self.auto_backup.get()))
            self.db.ayar_ekle("backup_interval", self.backup_interval.get())
            self.db.ayar_ekle("email_sender", self.email_sender.get())
            self.db.ayar_ekle("email_password", self.email_password.get())
            self.db.ayar_ekle("email_receiver", self.email_receiver.get())
            messagebox.showinfo("Başarılı", "Ayarlar kaydedildi")
        except Exception as e:
            messagebox.showerror("Hata", f"Ayarlar kaydedilemedi: {str(e)}")
            
    def show_about(self):
        """Hakkında penceresini gösterir"""
        messagebox.showinfo("Hakkında", 
            "Tamir Atölyesi Yönetim Sistemi\n"
            "Sürüm 1.0\n\n"
            "© 2024 Tüm hakları saklıdır.")
            
    def show_customer_report(self):
        """Müşteri raporunu gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Müşteri Raporu")
        dialog.geometry("800x600")
        # Tarih aralığı seçimi
        date_frame = ttk.LabelFrame(dialog, text="Tarih Aralığı", padding="10")
        date_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(date_frame, text="Başlangıç:").grid(row=0, column=0, padx=5, pady=5)
        baslangic = ttk.Entry(date_frame)
        baslangic.grid(row=0, column=1, padx=5, pady=5)
        baslangic.insert(0, datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(date_frame, text="Bitiş:").grid(row=0, column=2, padx=5, pady=5)
        bitis = ttk.Entry(date_frame)
        bitis.grid(row=0, column=3, padx=5, pady=5)
        bitis.insert(0, datetime.now().strftime("%d.%m.%Y"))
        # Rapor tablosu
        table_frame = ttk.Frame(dialog)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        columns = ("musteri", "tamir_sayisi", "toplam_ucret")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        tree.heading("musteri", text="Müşteri")
        tree.heading("tamir_sayisi", text="Tamir Sayısı")
        tree.heading("toplam_ucret", text="Toplam Ücret")
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        def rapor_olustur():
            try:
                for item in tree.get_children():
                    tree.delete(item)
                # Tarihleri veritabanı formatına çevir
                bas = datetime.strptime(baslangic.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                bit = datetime.strptime(bitis.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                veriler = self.db.musteri_tamir_raporu(bas, bit)
                for veri in veriler:
                    tree.insert("", "end", values=veri)
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        ttk.Button(date_frame, text="Rapor Oluştur", command=rapor_olustur).grid(row=0, column=4, padx=20, pady=5)
        
    def show_repair_report(self):
        """Tamir raporunu gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Tamir Raporu")
        dialog.geometry("800x600")
        # Tarih aralığı seçimi
        date_frame = ttk.LabelFrame(dialog, text="Tarih Aralığı", padding="10")
        date_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(date_frame, text="Başlangıç:").grid(row=0, column=0, padx=5, pady=5)
        baslangic = ttk.Entry(date_frame)
        baslangic.grid(row=0, column=1, padx=5, pady=5)
        baslangic.insert(0, datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(date_frame, text="Bitiş:").grid(row=0, column=2, padx=5, pady=5)
        bitis = ttk.Entry(date_frame)
        bitis.grid(row=0, column=3, padx=5, pady=5)
        bitis.insert(0, datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(date_frame, text="Durum:").grid(row=0, column=4, padx=5, pady=5)
        durum = ttk.Combobox(date_frame, values=["Tümü", "Beklemede", "Devam Ediyor", "Tamamlandı"])
        durum.grid(row=0, column=5, padx=5, pady=5)
        durum.set("Tümü")
        table_frame = ttk.Frame(dialog)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        columns = ("tarih", "musteri", "arac", "sorun", "durum", "ucret")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        tree.heading("tarih", text="Tarih")
        tree.heading("musteri", text="Müşteri")
        tree.heading("arac", text="Araç")
        tree.heading("sorun", text="Sorun")
        tree.heading("durum", text="Durum")
        tree.heading("ucret", text="Ücret")
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        def rapor_olustur():
            try:
                for item in tree.get_children():
                    tree.delete(item)
                # Tarihleri veritabanı formatına çevir
                bas = datetime.strptime(baslangic.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                bit = datetime.strptime(bitis.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                veriler = self.db.tamir_raporu(bas, bit, None if durum.get() == "Tümü" else durum.get())
                for veri in veriler:
                    tree.insert("", "end", values=veri)
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        ttk.Button(date_frame, text="Rapor Oluştur", command=rapor_olustur).grid(row=0, column=6, padx=20, pady=5)
        
    def show_income_report(self):
        """Gelir raporunu gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Gelir Raporu")
        dialog.geometry("600x400")
        date_frame = ttk.LabelFrame(dialog, text="Tarih Aralığı", padding="10")
        date_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(date_frame, text="Başlangıç:").grid(row=0, column=0, padx=5, pady=5)
        baslangic = ttk.Entry(date_frame)
        baslangic.grid(row=0, column=1, padx=5, pady=5)
        baslangic.insert(0, datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(date_frame, text="Bitiş:").grid(row=0, column=2, padx=5, pady=5)
        bitis = ttk.Entry(date_frame)
        bitis.grid(row=0, column=3, padx=5, pady=5)
        bitis.insert(0, datetime.now().strftime("%d.%m.%Y"))
        summary_frame = ttk.LabelFrame(dialog, text="Gelir Özeti", padding="10")
        summary_frame.pack(fill="x", padx=10, pady=10)
        ttk.Label(summary_frame, text="Toplam İşçilik:").grid(row=0, column=0, sticky="w", pady=5)
        toplam_iscilik = ttk.Label(summary_frame, text="0 TL")
        toplam_iscilik.grid(row=0, column=1, sticky="w", pady=5)
        ttk.Label(summary_frame, text="Toplam Parça:").grid(row=1, column=0, sticky="w", pady=5)
        toplam_parca = ttk.Label(summary_frame, text="0 TL")
        toplam_parca.grid(row=1, column=1, sticky="w", pady=5)
        ttk.Label(summary_frame, text="Genel Toplam:").grid(row=2, column=0, sticky="w", pady=5)
        genel_toplam = ttk.Label(summary_frame, text="0 TL")
        genel_toplam.grid(row=2, column=1, sticky="w", pady=5)
        def rapor_olustur():
            try:
                # Tarihleri veritabanı formatına çevir
                bas = datetime.strptime(baslangic.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                bit = datetime.strptime(bitis.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
                veriler = self.db.gelir_raporu(bas, bit)
                if veriler:
                    toplam_iscilik.config(text=f"{veriler['toplam_iscilik']:.2f} TL")
                    toplam_parca.config(text=f"{veriler['toplam_parca']:.2f} TL")
                    genel_toplam.config(text=f"{veriler['genel_toplam']:.2f} TL")
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        ttk.Button(date_frame, text="Rapor Oluştur", command=rapor_olustur).grid(row=0, column=4, padx=20, pady=5)
        
    def show_monthly_stats(self):
        """Aylık istatistikleri gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Aylık İstatistikler")
        dialog.geometry("600x400")
        
        # Ay seçimi
        month_frame = ttk.LabelFrame(dialog, text="Ay Seçimi", padding="10")
        month_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(month_frame, text="Ay:").grid(row=0, column=0, padx=5, pady=5)
        ay = ttk.Combobox(month_frame, values=[str(i) for i in range(1, 13)])
        ay.grid(row=0, column=1, padx=5, pady=5)
        ay.set(str(datetime.now().month))
        
        ttk.Label(month_frame, text="Yıl:").grid(row=0, column=2, padx=5, pady=5)
        yil = ttk.Entry(month_frame)
        yil.grid(row=0, column=3, padx=5, pady=5)
        yil.insert(0, str(datetime.now().year))
        
        # İstatistik özeti
        stats_frame = ttk.LabelFrame(dialog, text="İstatistikler", padding="10")
        stats_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(stats_frame, text="Toplam Tamir:").grid(row=0, column=0, sticky="w", pady=5)
        toplam_tamir = ttk.Label(stats_frame, text="0")
        toplam_tamir.grid(row=0, column=1, sticky="w", pady=5)
        
        ttk.Label(stats_frame, text="Ortalama Süre:").grid(row=1, column=0, sticky="w", pady=5)
        ortalama_sure = ttk.Label(stats_frame, text="0 gün")
        ortalama_sure.grid(row=1, column=1, sticky="w", pady=5)
        
        ttk.Label(stats_frame, text="Toplam Gelir:").grid(row=2, column=0, sticky="w", pady=5)
        toplam_gelir = ttk.Label(stats_frame, text="0 TL")
        toplam_gelir.grid(row=2, column=1, sticky="w", pady=5)
        
        def istatistik_goster():
            try:
                # İstatistikleri getir
                veriler = self.db.aylik_rapor(
                    int(ay.get()),
                    int(yil.get())
                )
                
                # Verileri göster
                if veriler:
                    toplam_tamir.config(text=str(veriler['toplam_tamir']))
                    ortalama_sure.config(text=f"{veriler['ortalama_sure']:.1f} gün")
                    toplam_gelir.config(text=f"{veriler['toplam_gelir']:.2f} TL")
                    
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        
        ttk.Button(month_frame, text="Göster", command=istatistik_goster).grid(row=0, column=4, padx=20, pady=5)
        
    def show_yearly_stats(self):
        """Yıllık istatistikleri gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Yıllık İstatistikler")
        dialog.geometry("800x600")
        
        # Yıl seçimi
        year_frame = ttk.LabelFrame(dialog, text="Yıl Seçimi", padding="10")
        year_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(year_frame, text="Yıl:").grid(row=0, column=0, padx=5, pady=5)
        yil = ttk.Entry(year_frame)
        yil.grid(row=0, column=1, padx=5, pady=5)
        yil.insert(0, str(datetime.now().year))
        
        # İstatistik tablosu
        table_frame = ttk.Frame(dialog)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        columns = ("ay", "tamir_sayisi", "ortalama_sure", "toplam_gelir")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        
        tree.heading("ay", text="Ay")
        tree.heading("tamir_sayisi", text="Tamir Sayısı")
        tree.heading("ortalama_sure", text="Ortalama Süre (Gün)")
        tree.heading("toplam_gelir", text="Toplam Gelir (TL)")
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def istatistik_goster():
            try:
                # Mevcut verileri temizle
                for item in tree.get_children():
                    tree.delete(item)
                    
                # İstatistikleri getir
                veriler = self.db.yillik_rapor(int(yil.get()))
                
                # Verileri tabloya ekle
                aylar = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
                
                for i, veri in enumerate(veriler):
                    tree.insert("", "end", values=(
                        aylar[i],
                        veri['tamir_sayisi'],
                        f"{veri['ortalama_sure']:.1f}",
                        f"{veri['toplam_gelir']:.2f}"
                    ))
                    
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        
        ttk.Button(year_frame, text="Göster", command=istatistik_goster).grid(row=0, column=2, padx=20, pady=5)
            
    def check_auto_backup(self):
        """Otomatik yedekleme kontrolü"""
        try:
            # Ayarları kontrol et
            auto_backup = self.db.ayar_getir("auto_backup")
            if auto_backup and auto_backup.lower() == "true":
                backup_interval = int(self.db.ayar_getir("backup_interval") or 7)
                
                # Son yedeği kontrol et
                yedekler = self.db.tum_yedeklemeler()
                if yedekler:
                    son_yedek = datetime.strptime(yedekler[0]['tarih'], "%Y-%m-%d %H:%M:%S")
                    gecen_gun = (datetime.now() - son_yedek).days
                    
                    if gecen_gun >= backup_interval:
                        # Yeni yedek al
                        self.db.yedek_al("Otomatik yedekleme")
                        logger.info("Otomatik yedek alındı")
                else:
                    # İlk yedek
                    self.db.yedek_al("İlk otomatik yedekleme")
                    logger.info("İlk otomatik yedek alındı")
                    
            # 24 saat sonra tekrar kontrol et
            self.parent.after(24 * 60 * 60 * 1000, self.check_auto_backup)
            
        except Exception as e:
            logger.error(f"Otomatik yedekleme hatası: {str(e)}")
            
    def __del__(self):
        """Yıkıcı metod"""
        pass

    def new_customer(self, edit_mode=False, musteri=None):
        """Yeni müşteri ekleme/düzenleme penceresi"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Müşteri Düzenle" if edit_mode else "Yeni Müşteri")
        dialog.geometry("400x500")
        dialog.resizable(False, False)
        
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Form alanları
        ttk.Label(main_frame, text="TC No:").pack(anchor="w", pady=(0, 5))
        tc_no = ttk.Entry(main_frame)
        tc_no.pack(fill="x", pady=(0, 10))
        
        ttk.Label(main_frame, text="Ad:").pack(anchor="w", pady=(0, 5))
        ad = ttk.Entry(main_frame)
        ad.pack(fill="x", pady=(0, 10))
        
        ttk.Label(main_frame, text="Soyad:").pack(anchor="w", pady=(0, 5))
        soyad = ttk.Entry(main_frame)
        soyad.pack(fill="x", pady=(0, 10))
        
        ttk.Label(main_frame, text="Telefon:").pack(anchor="w", pady=(0, 5))
        telefon = ttk.Entry(main_frame)
        telefon.pack(fill="x", pady=(0, 10))
        
        ttk.Label(main_frame, text="E-posta:").pack(anchor="w", pady=(0, 5))
        email = ttk.Entry(main_frame)
        email.pack(fill="x", pady=(0, 10))
        
        ttk.Label(main_frame, text="Adres:").pack(anchor="w", pady=(0, 5))
        adres = tk.Text(main_frame, height=4)
        adres.pack(fill="x", pady=(0, 10))
        
        # Düzenleme modunda alanları doldur
        if edit_mode and musteri:
            tc_no.insert(0, musteri[1] or "")
            ad.insert(0, musteri[2] or "")
            soyad.insert(0, musteri[3] or "")
            telefon.insert(0, musteri[4] or "")
            email.insert(0, musteri[5] or "")
            adres.insert("1.0", musteri[6] or "")
        
        def kaydet():
            try:
                # Zorunlu alan kontrolleri
                if not ad.get().strip() or not soyad.get().strip() or not telefon.get().strip():
                    messagebox.showwarning("Uyarı", "Lütfen ad, soyad ve telefon alanlarını doldurun!")
                    return
                if edit_mode:
                    if not musteri:
                        messagebox.showerror("Hata", "Düzenlenecek müşteri bulunamadı!")
                        return
                    musteri_id = musteri[0] if musteri else None
                    if musteri_id is None:
                        messagebox.showerror("Hata", "Müşteri ID'si alınamadı!")
                        return
                    # Müşteriyi güncelle
                    self.db.musteri_guncelle(
                        musteri_id,
                        tc_no.get(),
                        ad.get(),
                        soyad.get(),
                        telefon.get(),
                        email.get(),
                        adres.get("1.0", "end-1c")
                    )
                    messagebox.showinfo("Başarılı", "Müşteri bilgileri güncellendi")
                else:
                    # Yeni müşteri ekle
                    self.db.musteri_ekle(
                        tc_no.get(),
                        ad.get(),
                        soyad.get(),
                        telefon.get(),
                        email.get(),
                        adres.get("1.0", "end-1c")
                    )
                    messagebox.showinfo("Başarılı", "Yeni müşteri eklendi")
                
                # Müşteri listesini güncelle
                self.refresh_customers()
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        
        # Butonlar
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        ttk.Button(
            button_frame,
            text="Kaydet",
            command=kaydet
        ).pack(side="left", padx=5)
        
        ttk.Button(
            button_frame,
            text="İptal",
            command=dialog.destroy
        ).pack(side="left", padx=5)

    def new_repair(self, edit_mode=False, tamir=None):
        """Yeni tamir ekleme/düzenleme penceresi"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Yeni Tamir" if not edit_mode else "Tamir Düzenle")
        dialog.geometry("800x900")  # Yüksekliği artırdım
        dialog.minsize(800, 900)    # Minimum boyutu ayarladım
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Ana frame
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Form içeriği frame
        form_frame = ttk.Frame(main_frame)
        form_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # Müşteri seçimi
        customer_frame = ttk.LabelFrame(form_frame, text="Müşteri Bilgileri", padding="10")
        customer_frame.pack(fill="x", pady=5)
        
        ttk.Label(customer_frame, text="Müşteri:").grid(row=0, column=0, sticky="w", padx=5)
        self.customer_combo = ttk.Combobox(customer_frame, width=40)
        self.customer_combo.grid(row=0, column=1, sticky="ew", padx=5)
        
        # Cihaz bilgileri
        device_frame = ttk.LabelFrame(form_frame, text="Cihaz Bilgileri", padding="10")
        device_frame.pack(fill="x", pady=5)
        
        ttk.Label(device_frame, text="Cihaz Türü:").grid(row=0, column=0, sticky="w", padx=5)
        self.device_type = ttk.Combobox(device_frame, values=list(self.cihaz_kategorileri.keys()))
        self.device_type.grid(row=0, column=1, sticky="ew", padx=5)
        
        ttk.Label(device_frame, text="Marka:").grid(row=1, column=0, sticky="w", padx=5)
        self.brand = ttk.Entry(device_frame)
        self.brand.grid(row=1, column=1, sticky="ew", padx=5)
        
        ttk.Label(device_frame, text="Model:").grid(row=2, column=0, sticky="w", padx=5)
        self.model = ttk.Entry(device_frame)
        self.model.grid(row=2, column=1, sticky="ew", padx=5)
        
        ttk.Label(device_frame, text="Seri No:").grid(row=3, column=0, sticky="w", padx=5)
        self.serial = ttk.Entry(device_frame)
        self.serial.grid(row=3, column=1, sticky="ew", padx=5)
        
        # Arıza bilgileri
        problem_frame = ttk.LabelFrame(form_frame, text="Arıza Bilgileri", padding="10")
        problem_frame.pack(fill="x", pady=5)
        
        ttk.Label(problem_frame, text="Şikayet:").grid(row=0, column=0, sticky="w", padx=5)
        self.complaint = tk.Text(problem_frame, height=3)
        self.complaint.grid(row=0, column=1, sticky="ew", padx=5)
        
        ttk.Label(problem_frame, text="Tespit:").grid(row=1, column=0, sticky="w", padx=5)
        self.problem = tk.Text(problem_frame, height=3)
        self.problem.grid(row=1, column=1, sticky="ew", padx=5)
        
        ttk.Label(problem_frame, text="Yapılan İşlemler:").grid(row=2, column=0, sticky="w", padx=5)
        self.operations = tk.Text(problem_frame, height=3)
        self.operations.grid(row=2, column=1, sticky="ew", padx=5)
        
        # Grid yapılandırması
        problem_frame.grid_columnconfigure(1, weight=1)
        
        # Parça bilgileri
        parts_frame = ttk.LabelFrame(form_frame, text="Parça Bilgileri", padding="10")
        parts_frame.pack(fill="x", pady=5)
        
        # Parça tablosu
        self.parts_tree = ttk.Treeview(parts_frame, columns=("ad", "adet", "fiyat"), show="headings", height=5)
        self.parts_tree.heading("ad", text="Parça Adı")
        self.parts_tree.heading("adet", text="Adet")
        self.parts_tree.heading("fiyat", text="Fiyat")
        
        # Sütun genişlikleri
        self.parts_tree.column("ad", width=300)
        self.parts_tree.column("adet", width=100)
        self.parts_tree.column("fiyat", width=100)
        
        self.parts_tree.pack(fill="x", pady=5)
        
        # Parça butonları
        parts_button_frame = ttk.Frame(parts_frame)
        parts_button_frame.pack(fill="x")
        
        ttk.Button(parts_button_frame, text="Parça Ekle", command=self.add_part).pack(side="left", padx=5)
        ttk.Button(parts_button_frame, text="Parça Sil", command=self.remove_part).pack(side="left", padx=5)
        
        # Ücret bilgileri
        price_frame = ttk.LabelFrame(form_frame, text="Ücret Bilgileri", padding="10")
        price_frame.pack(fill="x", pady=5)
        
        ttk.Label(price_frame, text="İşçilik Ücreti:").grid(row=0, column=0, sticky="w", padx=5)
        self.labor_price = ttk.Entry(price_frame)
        self.labor_price.grid(row=0, column=1, sticky="ew", padx=5)
        self.labor_price.bind("<KeyRelease>", lambda e: self.calculate_total())
        
        ttk.Label(price_frame, text="Parça Ücreti:").grid(row=1, column=0, sticky="w", padx=5)
        self.parts_price = ttk.Entry(price_frame)
        self.parts_price.grid(row=1, column=1, sticky="ew", padx=5)
        self.parts_price.bind("<KeyRelease>", lambda e: self.calculate_total())
        
        ttk.Label(price_frame, text="Toplam Ücret:").grid(row=2, column=0, sticky="w", padx=5)
        self.total_price = ttk.Entry(price_frame, state="readonly")
        self.total_price.grid(row=2, column=1, sticky="ew", padx=5)
        
        # Grid yapılandırması
        price_frame.grid_columnconfigure(1, weight=1)
        
        # Durum seçimi
        status_frame = ttk.LabelFrame(form_frame, text="Durum Bilgisi", padding="10")
        status_frame.pack(fill="x", pady=5)
        
        ttk.Label(status_frame, text="Durum:").pack(side="left", padx=5)
        self.status = ttk.Combobox(status_frame, values=["Beklemede", "İncelemede", "Tamir Ediliyor", "Tamamlandı"])
        self.status.pack(side="left", padx=5)
        self.status.set("Beklemede")
        
        # Alt buton çerçevesi
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))
        
        # Kaydet ve İptal butonları
        ttk.Button(
            button_frame,
            text="Kaydet",
            command=self.save_repair,
            style="Accent.TButton"  # Özel stil ekledim
        ).pack(side="right", padx=5)
        
        ttk.Button(
            button_frame,
            text="İptal",
            command=dialog.destroy
        ).pack(side="right", padx=5)
        
        # Müşteri listesini yükle
        self.load_customers()
        
        # Düzenleme modunda verileri yükle
        if edit_mode and tamir:
            self.load_repair_data(tamir)
            
    def add_part(self):
        """Parça ekleme penceresi"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Parça Ekle")
        dialog.geometry("400x200")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        # Ana frame
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Form alanları
        ttk.Label(main_frame, text="Parça Adı:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        parca_adi = ttk.Entry(main_frame, width=30)
        parca_adi.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(main_frame, text="Adet:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        adet = ttk.Entry(main_frame, width=10)
        adet.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        adet.insert(0, "1")
        
        ttk.Label(main_frame, text="Fiyat:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        fiyat = ttk.Entry(main_frame, width=15)
        fiyat.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        fiyat.insert(0, "0.00")
        
        def kaydet():
            try:
                if not parca_adi.get().strip():
                    messagebox.showwarning("Uyarı", "Lütfen parça adı girin!")
                    return
                    
                adet_value = int(adet.get())
                fiyat_value = float(fiyat.get())
                
                self.parts_tree.insert("", "end", values=(
                    parca_adi.get().strip(),
                    str(adet_value),
                    "{:.2f} TL".format(fiyat_value)
                ))
                
                self.calculate_total()
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("Hata", "Lütfen geçerli sayısal değerler girin!")
        
        # Butonlar
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Kaydet", command=kaydet).pack(side="left", padx=5)
        ttk.Button(button_frame, text="İptal", command=dialog.destroy).pack(side="left", padx=5)

    def calculate_total(self):
        """Toplam ücreti hesaplar"""
        try:
            iscilik = float(self.labor_price.get() or 0)
            parca = float(self.parts_price.get() or 0)
            toplam = iscilik + parca
            
            self.total_price.configure(state="normal")
            self.total_price.delete(0, "end")
            self.total_price.insert(0, f"{toplam:.2f}")
            self.total_price.configure(state="readonly")
        except:
            pass
            
    def load_customers(self):
        """Müşteri listesini yükler"""
        try:
            musteriler = self.db.tum_musteriler()
            self.musteri_listesi = {f"{m['ad']} {m['soyad']}": m['id'] for m in musteriler}
            self.customer_combo["values"] = list(self.musteri_listesi.keys())
        except Exception as e:
            messagebox.showerror("Hata", f"Müşteri listesi yüklenirken hata oluştu: {str(e)}")

    def setup_dashboard_tab(self):
        """Ana sayfa sekmesini hazırlar"""
        # Ana frame
        self.dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.dashboard_frame, text="Ana Sayfa")

        # İstatistik kartları için frame
        stats_frame = ttk.LabelFrame(self.dashboard_frame, text="Genel İstatistikler", padding="10")
        stats_frame.pack(fill="x", padx=10, pady=5)

        # İstatistik kartları
        self.total_customers_label = ttk.Label(stats_frame, text="Toplam Müşteri: 0")
        self.total_customers_label.grid(row=0, column=0, padx=10, pady=5)

        self.active_repairs_label = ttk.Label(stats_frame, text="Aktif Tamir: 0")
        self.active_repairs_label.grid(row=0, column=1, padx=10, pady=5)

        self.completed_repairs_label = ttk.Label(stats_frame, text="Tamamlanan Tamir: 0")
        self.completed_repairs_label.grid(row=0, column=2, padx=10, pady=5)

        self.monthly_income_label = ttk.Label(stats_frame, text="Aylık Gelir: 0 TL")
        self.monthly_income_label.grid(row=1, column=0, padx=10, pady=5)

        self.yearly_income_label = ttk.Label(stats_frame, text="Yıllık Gelir: 0 TL")
        self.yearly_income_label.grid(row=1, column=1, padx=10, pady=5)

        # Son işlemler için frame
        recent_frame = ttk.LabelFrame(self.dashboard_frame, text="Son İşlemler", padding="10")
        recent_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Son tamirler listesi
        self.recent_repairs_tree = ttk.Treeview(
            recent_frame,
            columns=("id", "tarih", "musteri", "cihaz", "durum", "ucret"),
            show="headings",
            height=10
        )
        self.recent_repairs_tree.heading("id", text="ID")
        self.recent_repairs_tree.heading("tarih", text="Tarih")
        self.recent_repairs_tree.heading("musteri", text="Müşteri")
        self.recent_repairs_tree.heading("cihaz", text="Cihaz")
        self.recent_repairs_tree.heading("durum", text="Durum")
        self.recent_repairs_tree.heading("ucret", text="Ücret")
        self.recent_repairs_tree.column("id", width=50)
        self.recent_repairs_tree.column("tarih", width=100)
        self.recent_repairs_tree.column("musteri", width=150)
        self.recent_repairs_tree.column("cihaz", width=150)
        self.recent_repairs_tree.column("durum", width=100)
        self.recent_repairs_tree.column("ucret", width=100)
        self.recent_repairs_tree.pack(fill="both", expand=True)
        # Scrollbar
        scrollbar = ttk.Scrollbar(recent_frame, orient="vertical", command=self.recent_repairs_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.recent_repairs_tree.configure(yscrollcommand=scrollbar.set)
        # İstatistikleri güncelle
        self.update_dashboard()

    def update_dashboard(self):
        """Ana sayfa istatistiklerini günceller"""
        try:
            # Tamir istatistikleri
            tamirler = self.db.tum_tamirler()
            bekleyen = 0
            devam_eden = 0
            tamamlanan = 0
            toplam_gelir = 0.0
            
            # Son işlemleri temizle
            for item in self.recent_repairs_tree.get_children():
                self.recent_repairs_tree.delete(item)
            
            # Tamirler üzerinde döngü
            for tamir in tamirler:
                try:
                    # Durum kontrolü
                    durum = str(tamir[5]) if tamir[5] else ""
                    if durum == "Beklemede":
                        bekleyen += 1
                    elif durum == "Tamamlandı":
                        tamamlanan += 1
                    else:
                        devam_eden += 1
                    
                    # Ücret hesaplama
                    ucret = 0.0
                    if tamir[6]:
                        try:
                            ucret = float(tamir[6])
                            toplam_gelir += ucret
                        except:
                            pass
                    
                    # Son 10 işlemi listele
                    if len(self.recent_repairs_tree.get_children()) < 10:
                        musteri_adi = f"{tamir[2]} {tamir[3]}" if tamir[2] and tamir[3] else ""
                        cihaz = str(tamir[4]) if tamir[4] else ""
                        
                        self.recent_repairs_tree.insert("", "end", values=(
                            str(tamir[0]) if tamir[0] else "",  # ID
                            str(tamir[1]) if tamir[1] else "",  # Tarih
                            musteri_adi,  # Müşteri
                            cihaz,  # Cihaz
                            durum,  # Durum
                            f"{ucret:.2f} TL"  # Ücret
                        ))
                        
                except Exception as e:
                    logger.error(f"Tamir kaydı işlenirken hata: {str(e)}")
                    continue
            
            # İstatistikleri güncelle
            self.total_customers_label.config(text=f"Toplam Müşteri: {self.db.tum_musteri_sayisi()}")
            self.active_repairs_label.config(text=f"Aktif Tamir: {devam_eden}")
            self.completed_repairs_label.config(text=f"Tamamlanan Tamir: {tamamlanan}")
            self.monthly_income_label.config(text=f"Aylık Gelir: {toplam_gelir:.2f} TL")
            
            # Yıllık geliri ayrı hesapla
            try:
                yillik_gelir = self.db.yillik_gelir()
                self.yearly_income_label.config(text=f"Yıllık Gelir: {yillik_gelir:.2f} TL")
            except:
                self.yearly_income_label.config(text="Yıllık Gelir: 0.00 TL")
                
        except Exception as e:
            logger.error(f"Dashboard güncellenirken hata: {str(e)}")
            messagebox.showerror("Hata", "İstatistikler güncellenirken bir hata oluştu.")

    def generate_report(self):
        """Seçili raporu oluşturur ve özet, tablo, grafik alanlarını günceller"""
        rapor_turu = self.report_type.get()
        # Tarihleri veritabanı formatına çevir
        baslangic = datetime.strptime(self.start_date.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
        bitis = datetime.strptime(self.end_date.get(), "%d.%m.%Y").strftime("%Y-%m-%d")
        durum = self.status_filter.get() if rapor_turu == "Tamir Raporu" else None
        # Tablo ve özet kutucuklarını temizle
        for lbl in self.summary_labels:
            lbl.destroy()
        self.summary_labels = []
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        # Grafik alanını temizle
        for widget in self.report_graph_frame.winfo_children():
            widget.destroy()
        self.report_graph_canvas = None
        # --- Rapor türüne göre işlemler ---
        if rapor_turu == "Tamir Raporu":
            self.report_tree["columns"] = ("giris_tarihi", "musteri", "cihaz", "durum", "ucret")
            for col, text, w in zip(self.report_tree["columns"], ["Giriş Tarihi", "Müşteri", "Cihaz", "Durum", "Ücret"], [120, 200, 200, 100, 100]):
                self.report_tree.heading(col, text=text)
                self.report_tree.column(col, width=w)
            veriler = self.db.tamir_raporu(baslangic, bitis, None if durum == "Tümü" else durum)
            toplam_ucret = 0
            for veri in veriler:
                self.report_tree.insert("", "end", values=(veri['giris_tarihi'], veri['musteri_adi'], f"{veri['cihaz_turu']} {veri['marka']} {veri['model']}", veri['durum'], f"{veri['toplam_ucret']:.2f} TL"))
                try:
                    toplam_ucret += float(veri['toplam_ucret'])
                except:
                    pass
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Kayıt: {len(veriler)}", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Ücret: {toplam_ucret:.2f} TL", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            # Grafik: Durumlara göre dağılım
            from collections import Counter
            durumlar = [v['durum'] for v in veriler]
            if durumlar:
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                fig, ax = plt.subplots(figsize=(3,3))
                sayim = Counter(durumlar)
                ax.pie(list(sayim.values()), labels=list(sayim.keys()), autopct='%1.1f%%', startangle=90)
                ax.set_title('Durum Dağılımı')
                self.report_graph_canvas = FigureCanvasTkAgg(fig, master=self.report_graph_frame)
                self.report_graph_canvas.draw()
                self.report_graph_canvas.get_tk_widget().pack(fill="both", expand=True)
        elif rapor_turu == "Müşteri Raporu":
            self.report_tree["columns"] = ("musteri", "tamir_sayisi", "toplam_ucret", "son_tamir")
            for col, text, w in zip(self.report_tree["columns"], ["Müşteri", "Tamir Sayısı", "Toplam Ücret", "Son Tamir"], [200, 100, 120, 120]):
                self.report_tree.heading(col, text=text)
                self.report_tree.column(col, width=w)
            musteriler = self.db.musteri_bazli_tamir_raporu(baslangic, bitis)
            toplam_tamir = 0
            toplam_ucret = 0
            for musteri in musteriler:
                self.report_tree.insert("", "end", values=(f"{musteri[2]} {musteri[3]}", musteri[8], f"{musteri[11]:.2f} TL", musteri[7]))
                toplam_tamir += musteri[8]
                toplam_ucret += musteri[11]
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Müşteri: {len(musteriler)}", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Tamir: {toplam_tamir}", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Ücret: {toplam_ucret:.2f} TL", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            # Grafik: Tamir sayısına göre ilk 5 müşteri
            if musteriler:
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                ilk5 = sorted(musteriler, key=lambda x: x[8], reverse=True)[:5]
                fig, ax = plt.subplots(figsize=(4,3))
                ax.bar([f"{m[2]} {m[3]}" for m in ilk5], [m[8] for m in ilk5], color='skyblue')
                ax.set_title('En Çok Tamir Yapan 5 Müşteri')
                ax.set_ylabel('Tamir Sayısı')
                self.report_graph_canvas = FigureCanvasTkAgg(fig, master=self.report_graph_frame)
                self.report_graph_canvas.draw()
                self.report_graph_canvas.get_tk_widget().pack(fill="both", expand=True)
        elif rapor_turu == "Gelir Raporu":
            self.report_tree["columns"] = ("kategori", "miktar")
            for col, text, w in zip(self.report_tree["columns"], ["Kategori", "Miktar"], [200, 120]):
                self.report_tree.heading(col, text=text)
                self.report_tree.column(col, width=w)
            gelir = self.db.gelir_raporu(baslangic, bitis)
            if gelir:
                self.report_tree.insert("", "end", values=("İşçilik Geliri", f"{gelir['toplam_iscilik']:.2f} TL"))
                self.report_tree.insert("", "end", values=("Parça Geliri", f"{gelir['toplam_parca']:.2f} TL"))
                self.report_tree.insert("", "end", values=("Toplam Gelir", f"{gelir['genel_toplam']:.2f} TL"))
                self.report_tree.insert("", "end", values=("Tamir Sayısı", gelir['tamir_sayisi']))
                self.report_tree.insert("", "end", values=("Ortalama Ücret", f"{gelir['ortalama_ucret']:.2f} TL"))
                self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Gelir: {gelir['genel_toplam']:.2f} TL", style="Accent.TLabel"))
                self.summary_labels[-1].pack(side="left", padx=10)
                self.summary_labels.append(ttk.Label(self.summary_frame, text=f"İşçilik: {gelir['toplam_iscilik']:.2f} TL", style="Accent.TLabel"))
                self.summary_labels[-1].pack(side="left", padx=10)
                self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Parça: {gelir['toplam_parca']:.2f} TL", style="Accent.TLabel"))
                self.summary_labels[-1].pack(side="left", padx=10)
                if gelir['toplam_iscilik'] > 0 or gelir['toplam_parca'] > 0:
                    import matplotlib.pyplot as plt
                    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                    fig, ax = plt.subplots(figsize=(3,3))
                    ax.pie([gelir['toplam_iscilik'], gelir['toplam_parca']], labels=["İşçilik", "Parça"], autopct='%1.1f%%', startangle=90)
                    ax.set_title('Gelir Dağılımı')
                    self.report_graph_canvas = FigureCanvasTkAgg(fig, master=self.report_graph_frame)
                    self.report_graph_canvas.draw()
                    self.report_graph_canvas.get_tk_widget().pack(fill="both", expand=True)
        elif rapor_turu == "Maliyet Analizi":
            self.report_tree["columns"] = ("cihaz", "iscilik", "parca", "toplam", "kar_orani")
            for col, text, w in zip(self.report_tree["columns"], ["Cihaz", "İşçilik", "Parça", "Toplam", "Kâr Oranı"], [200, 100, 100, 100, 100]):
                self.report_tree.heading(col, text=text)
                self.report_tree.column(col, width=w)
            maliyetler = self.db.maliyet_analiz_raporu(baslangic, bitis)
            toplam_iscilik = toplam_parca = toplam_toplam = 0
            for maliyet in maliyetler:
                self.report_tree.insert("", "end", values=(
                    maliyet[3],  # Cihaz
                    f"{maliyet[10]:.2f} TL",  # İşçilik
                    f"{maliyet[11]:.2f} TL",  # Parça
                    f"{maliyet[12]:.2f} TL",  # Toplam
                    f"%{((maliyet[12] - (maliyet[10] + maliyet[11])) / maliyet[12] * 100):.2f}" if maliyet[12] else "-"
                ))
                toplam_iscilik += maliyet[10]
                toplam_parca += maliyet[11]
                toplam_toplam += maliyet[12]
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Kayıt: {len(maliyetler)}", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam İşçilik: {toplam_iscilik:.2f} TL", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam Parça: {toplam_parca:.2f} TL", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            self.summary_labels.append(ttk.Label(self.summary_frame, text=f"Toplam: {toplam_toplam:.2f} TL", style="Accent.TLabel"))
            self.summary_labels[-1].pack(side="left", padx=10)
            if toplam_toplam > 0:
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                fig, ax = plt.subplots(figsize=(3,3))
                ax.pie([toplam_iscilik, toplam_parca], labels=["İşçilik", "Parça"], autopct='%1.1f%%', startangle=90)
                ax.set_title('Maliyet Dağılımı')
                self.report_graph_canvas = FigureCanvasTkAgg(fig, master=self.report_graph_frame)
                self.report_graph_canvas.draw()
                self.report_graph_canvas.get_tk_widget().pack(fill="both", expand=True)
        else:
            pass

    def refresh_backups(self):
        """Yedek listesini günceller"""
        for item in self.backup_tree.get_children():
            self.backup_tree.delete(item)
        yedekler = self.db.tum_yedeklemeler()
        print('Yedekler:', yedekler)  # DEBUG
        for yedek in yedekler:
            try:
                tarih = datetime.strptime(yedek["tarih"], "%Y-%m-%d %H:%M:%S") + timedelta(hours=3)
                tarih_str = tarih.strftime("%d.%m.%Y %H:%M:%S")
            except Exception:
                tarih_str = yedek["tarih"]
            boyut_kb = int(yedek["boyut"]) // 1024 if yedek["boyut"] else 0
            dosya_adi = yedek["dosya_adi"]
            dosya_yolu = os.path.join("backups", dosya_adi)
            exists = os.path.exists(dosya_yolu)
            row_id = self.backup_tree.insert(
                "", tk.END,
                values=(yedek["id"], tarih_str, dosya_adi, f"{boyut_kb} KB", yedek["aciklama"])
            )
            if not exists:
                self.backup_tree.item(row_id, tags=("eksik",))
        self.backup_tree.tag_configure("eksik", background="tomato")

    def create_backup(self):
        """Manuel yedek alma"""
        try:
            self.db.yedek_al("Manuel yedekleme")
            self.refresh_backups()
            self.refresh_backup_manager()  # Yedekleme yöneticisi de güncellensin
            yedekler = self.db.tum_yedeklemeler()
            if yedekler:
                dosya_adi = yedekler[0]["dosya_adi"]
                dosya_yolu = os.path.join("backups", dosya_adi)
                result = self.send_backup_email(dosya_yolu)
                if result:
                    messagebox.showinfo("Başarılı", "Yedek alındı ve e-posta gönderildi")
                else:
                    messagebox.showwarning("Uyarı", "Yedek alındı fakat e-posta gönderilemedi!")
            else:
                messagebox.showinfo("Başarılı", "Yedek alındı")
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def restore_backup(self):
        """Seçili yedeği geri yükle"""
        selected = self.backup_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen bir yedek seçin")
            return
        values = self.backup_tree.item(selected[0])["values"]
        yedek_id = values[0]
        dosya_adi = values[2]
        dosya_yolu = os.path.join("backups", dosya_adi)
        if not os.path.exists(dosya_yolu):
            messagebox.showerror("Hata", f"Yedek dosyası fiziksel olarak bulunamadı:\n{dosya_yolu}")
            return
        if messagebox.askyesno("Onay", "Seçili yedeği geri yüklemek istediğinize emin misiniz?"):
            try:
                self.db.yedek_geri_yukle(yedek_id)
                messagebox.showinfo("Başarılı", "Yedek geri yüklendi")
            except Exception as e:
                messagebox.showerror("Hata", str(e))

    def remove_part(self):
        """Seçili parçayı siler"""
        selected = self.parts_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen silinecek parçayı seçin!")
            return
            
        if messagebox.askyesno("Onay", "Seçili parçayı silmek istediğinize emin misiniz?"):
            self.parts_tree.delete(selected[0])
            self.calculate_total()
            
    def save_repair(self):
        """Tamir kaydını kaydeder"""
        try:
            # Müşteri kontrolü
            musteri = self.customer_combo.get()
            if not musteri or musteri not in self.musteri_listesi:
                messagebox.showwarning("Uyarı", "Lütfen müşteri seçin!")
                return
            
            # Zorunlu alanları kontrol et
            if not self.device_type.get():
                messagebox.showwarning("Uyarı", "Lütfen cihaz türünü seçin!")
                return
                
            if not self.complaint.get("1.0", "end-1c").strip():
                messagebox.showwarning("Uyarı", "Lütfen şikayet açıklamasını girin!")
                return
            
            # Parça listesini hazırla
            parcalar = []
            for item in self.parts_tree.get_children():
                values = self.parts_tree.item(item)["values"]
                parcalar.append({
                    "ad": values[0],
                    "adet": values[1],
                    "fiyat": float(values[2].replace(" TL", ""))
                })
            
            # Tamir kaydını ekle/güncelle
            if hasattr(self, 'edit_mode') and self.edit_mode:
                self.db.tamir_guncelle(
                    self.tamir_id,
                    self.musteri_listesi[musteri],
                    self.device_type.get(),
                    self.brand.get(),
                    self.model.get(),
                    self.serial.get(),
                    self.complaint.get("1.0", "end-1c"),
                    self.problem.get("1.0", "end-1c"),
                    self.operations.get("1.0", "end-1c"),
                    json.dumps(parcalar),
                    float(self.labor_price.get() or 0),
                    float(self.parts_price.get() or 0),
                    float(self.total_price.get().replace(" TL", "") or 0),
                    self.status.get()
                )
                messagebox.showinfo("Başarılı", "Tamir kaydı güncellendi")
            else:
                self.db.tamir_ekle(
                    self.musteri_listesi[musteri],
                    self.device_type.get(),
                    self.brand.get(),
                    self.model.get(),
                    self.serial.get(),
                    self.complaint.get("1.0", "end-1c"),
                    self.problem.get("1.0", "end-1c"),
                    self.operations.get("1.0", "end-1c"),
                    json.dumps(parcalar),
                    float(self.labor_price.get() or 0),
                    float(self.parts_price.get() or 0),
                    float(self.total_price.get().replace(" TL", "") or 0),
                    self.status.get()
                )
                messagebox.showinfo("Başarılı", "Yeni tamir kaydı eklendi")
            
            # Tamir listesini güncelle
            self.refresh_repairs()
            self.parent.focus_set()  # Ana pencereye odaklan
            
        except Exception as e:
            messagebox.showerror("Hata", str(e))
            
    def load_repair_data(self, tamir):
        """Tamir verilerini forma yükler"""
        try:
            tamir = dict(tamir)
            self.edit_mode = True
            self.tamir_id = tamir.get('id', '')
            self.customer_combo.set(f"{tamir.get('ad', '')} {tamir.get('soyad', '')}")
            self.device_type.set(tamir.get('cihaz_turu', ''))
            self.brand.insert(0, tamir.get('marka', ''))
            self.model.insert(0, tamir.get('model', ''))
            self.serial.insert(0, tamir.get('seri_no', ''))
            self.complaint.insert("1.0", tamir.get('sikayet', ''))
            self.problem.insert("1.0", tamir.get('ariza', ''))
            self.operations.insert("1.0", tamir.get('islemler', ''))
            self.labor_price.delete(0, "end")
            self.labor_price.insert(0, str(tamir.get('iscilik_ucreti', 0)))
            self.parts_price.delete(0, "end")
            self.parts_price.insert(0, str(tamir.get('parca_ucreti', 0)))
            self.total_price.configure(state="normal")
            self.total_price.delete(0, "end")
            self.total_price.insert(0, str(tamir.get('toplam_ucret', 0)))
            self.total_price.configure(state="readonly")
            self.status.set(tamir.get('durum', 'Beklemede'))
            # Parçaları yükle
            if tamir.get('parcalar'):
                try:
                    parcalar = json.loads(tamir['parcalar'])
                    for parca in parcalar:
                        self.parts_tree.insert("", "end", values=(
                            parca.get('ad', ''),
                            parca.get('adet', ''),
                            "{:.2f} TL".format(parca.get('fiyat', 0))
                        ))
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror("Hata", f"Tamir verileri yüklenirken hata oluştu: {str(e)}")

    def show_user_manual(self):
        """Geliştirilebilir kullanıcı kılavuzunu gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Kullanıcı Kılavuzu")
        dialog.geometry("600x500")
        dialog.transient(self.parent)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)

        # Kaydırılabilir metin kutusu
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill="both", expand=True)

        manual_text = (
            "Tamir Atölyesi Yönetim Sistemi Kullanıcı Kılavuzu\n"
            "-----------------------------------------------\n"
            "\n"
            "• Yeni müşteri eklemek için: 'Yeni Müşteri' butonunu veya menüsünü kullanın.\n"
            "• Tamir kaydı eklemek için: 'Yeni Tamir' butonunu veya menüsünü kullanın.\n"
            "• Yedekleme işlemleri için: Ayarlar sekmesinden veya 'Yedekleme Yöneticisi'nden işlem yapabilirsiniz.\n"
            "• Raporlar sekmesinden çeşitli raporları oluşturabilirsiniz.\n"
            "• Detaylı bilgi için geliştiriciye başvurun.\n"
            "\n"
            "Bu alanı dilediğiniz gibi genişletebilir, yeni başlıklar ve açıklamalar ekleyebilirsiniz.\n"
        )

        text_widget = tk.Text(text_frame, wrap="word")
        text_widget.insert("1.0", manual_text)
        text_widget.configure(state="disabled")
        text_widget.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        scrollbar.pack(side="right", fill="y")
        text_widget.configure(yscrollcommand=scrollbar.set)

        ttk.Button(main_frame, text="Kapat", command=dialog.destroy).pack(pady=10)

    def setup_stats_tab(self):
        """İstatistikler sekmesini hazırlar"""
        self.stats_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_tab, text="İstatistikler")

        # Grafik frame
        graph_frame = ttk.LabelFrame(self.stats_tab, text="Aylık Tamir ve Gelir Grafikleri", padding="10")
        graph_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Verileri çek
        now = datetime.now()
        aylik = self.db.aylik_rapor(now.month, now.year)
        yillik = self.db.yillik_rapor(now.year)
        if not aylik:
            aylik = {'toplam_iscilik': 0, 'toplam_parca': 0}
        if not yillik:
            yillik = [{'tamir_sayisi': 0} for _ in range(12)]
        # NaN kontrolü
        for k in ['toplam_iscilik', 'toplam_parca']:
            if k in aylik and (aylik[k] is None or (isinstance(aylik[k], float) and math.isnan(aylik[k]))):
                aylik[k] = 0
        for v in yillik:
            if 'tamir_sayisi' in v and (v['tamir_sayisi'] is None or (isinstance(v['tamir_sayisi'], float) and math.isnan(v['tamir_sayisi']))):
                v['tamir_sayisi'] = 0

        # Bar grafik: Yıllık tamir sayısı
        fig1, ax1 = plt.subplots(figsize=(5,3))
        aylar = ["Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
        tamirler = [v['tamir_sayisi'] for v in yillik]
        if any(tamirler):
            ax1.bar(aylar, tamirler, color='skyblue')
            ax1.set_title('Aylara Göre Tamir Sayısı')
            ax1.set_ylabel('Tamir')
        else:
            ax1.text(0.5, 0.5, 'Veri yok', ha='center', va='center')
        canvas1 = FigureCanvasTkAgg(fig1, master=graph_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(side="left", fill="both", expand=True, padx=10)

        # Pasta grafik: Aylık gelir dağılımı
        fig2, ax2 = plt.subplots(figsize=(4,4))
        labels = ['İşçilik', 'Parça']
        gelirler = [aylik['toplam_iscilik'], aylik['toplam_parca']]
        if sum(gelirler) > 0:
            ax2.pie(gelirler, labels=labels, autopct='%1.1f%%', startangle=90)
            ax2.set_title('Aylık Gelir Dağılımı')
        else:
            ax2.text(0.5, 0.5, 'Veri yok', ha='center', va='center')
        canvas2 = FigureCanvasTkAgg(fig2, master=graph_frame)
        canvas2.draw()
        canvas2.get_tk_widget().pack(side="right", fill="both", expand=True, padx=10)

    def export_report_excel(self):
        """Raporu Excel dosyasına aktarır"""
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
        if not file_path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        if ws:
            # Sütun başlıkları
            columns = self.report_tree["columns"]
            ws.append([self.report_tree.heading(col)["text"] for col in columns])
            # Satırlar
            for item in self.report_tree.get_children():
                values = self.report_tree.item(item)["values"]
                if isinstance(values, list) and values:
                    ws.append(values)
            # Sütun genişliklerini ayarla
            for i, col in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
            wb.save(file_path)

    def export_report_pdf(self):
        """Raporu PDF dosyasına aktarır (Türkçe karakter desteği ile)"""
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")])
        if not file_path:
            return
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        # Türkçe karakter desteği için DejaVuSans.ttf fontunu ekle
        pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
        pdf.set_font('DejaVu', '', 10)
        # Sütun başlıkları
        columns = self.report_tree["columns"]
        col_width = pdf.w / (len(columns) + 1)
        for col in columns:
            pdf.cell(col_width, 10, str(self.report_tree.heading(col)["text"]), border=1)
        pdf.ln()
        # Satırlar
        for item in self.report_tree.get_children():
            for value in self.report_tree.item(item)["values"]:
                pdf.cell(col_width, 10, str(value), border=1)
            pdf.ln()
        pdf.output(file_path)

    def load_email_settings(self):
        try:
            sender = self.db.ayar_getir("email_sender")
            if sender:
                self.email_sender.insert(0, sender)
            password = self.db.ayar_getir("email_password")
            if password:
                self.email_password.insert(0, password)
            receiver = self.db.ayar_getir("email_receiver")
            if receiver:
                self.email_receiver.insert(0, receiver)
        except Exception as e:
            logger.error(f"E-posta ayarları yüklenirken hata: {str(e)}")

    def send_backup_email(self, dosya_yolu=None):
        try:
            sender = self.email_sender.get()
            password = self.email_password.get()
            receiver = self.email_receiver.get()
            if not dosya_yolu:
                yedekler = self.db.tum_yedeklemeler()
                if not yedekler:
                    return False
                dosya_adi = yedekler[0]["dosya_adi"]
                dosya_yolu = os.path.join("backups", dosya_adi)
            if not os.path.exists(dosya_yolu):
                return False
            yag = yagmail.SMTP(sender, password)
            yag.send(
                to=receiver,
                subject="Tamir Atölyesi Yedek Dosyası",
                contents="Yedek dosyası ektedir.",
                attachments=dosya_yolu
            )
            return True
        except Exception as e:
            messagebox.showerror("E-posta Hatası", f"Yedek e-posta ile gönderilemedi: {str(e)}")
            return False

    def save_email_settings(self):
        try:
            self.db.ayar_ekle("email_sender", self.email_sender.get())
            self.db.ayar_ekle("email_password", self.email_password.get())
            self.db.ayar_ekle("email_receiver", self.email_receiver.get())
            messagebox.showinfo("Başarılı", "E-posta ayarları kaydedildi")
        except Exception as e:
            messagebox.showerror("Hata", f"E-posta ayarları kaydedilemedi: {str(e)}")

    def add_test_data(self):
        """Test verisi ekler (giris_tarihi: sistem saati)"""
        from datetime import datetime, timedelta
        try:
            # 5 müşteri ekle
            for i in range(1, 6):
                ad = f"TestAd{i}"
                soyad = f"TestSoyad{i}"
                telefon = f"55500000{i}"
                email = f"test{i}@mail.com"
                adres = f"Test Mah. {i}. Sokak No:{i}"
                self.db.musteri_ekle("", ad, soyad, telefon, email, adres)
            musteriler = self.db.tum_musteriler()
            cihazlar = ["Bilgisayar", "Telefon", "Televizyon", "Beyaz Eşya", "Yazıcı"]
            for idx, musteri in enumerate(musteriler[:5]):
                for j in range(2):
                    giris_tarihi = (datetime.utcnow() + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")
                    self.db.cursor.execute(
                        """
                        INSERT INTO tamirler (
                            musteri_id, cihaz_turu, marka, model, seri_no, sikayet, ariza, islemler, parcalar,
                            iscilik_ucreti, parca_ucreti, toplam_ucret, durum, giris_tarihi
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            musteri['id'],
                            cihazlar[(idx+j)%len(cihazlar)],
                            f"Marka{j+1}",
                            f"Model{j+1}",
                            f"SN{idx}{j}",
                            f"Şikayet örneği {j+1}",
                            f"Arıza örneği {j+1}",
                            f"İşlem örneği {j+1}",
                            "[]",
                            100.0 + 10*j,
                            50.0 + 5*j,
                            150.0 + 15*j,
                            "Tamamlandı" if j%2==0 else "Beklemede",
                            giris_tarihi
                        )
                    )
            self.db.conn.commit()
            self.refresh_customers()
            self.refresh_repairs()
            messagebox.showinfo("Başarılı", "Test verileri eklendi!")
        except Exception as e:
            messagebox.showerror("Hata", f"Test verisi eklenemedi: {str(e)}")

    def ekle_giris_tarihi_alani(self):
        try:
            self.db.giris_tarihi_alani_ekle()
            messagebox.showinfo("Başarılı", "giris_tarihi alanı başarıyla eklendi (veya zaten vardı). Artık test verisi ekleyebilirsiniz.")
        except Exception as e:
            messagebox.showerror("Hata", f"Alan eklenemedi: {str(e)}")

    def sil_test_tamirleri(self):
        try:
            self.db.test_tamirlerini_sil()
            self.refresh_customers()
            self.refresh_repairs()
            messagebox.showinfo("Başarılı", "Test verileri silindi!")
        except Exception as e:
            messagebox.showerror("Hata", f"Test verileri silinemedi: {str(e)}")

    def doldur_eksik_giris_tarihi(self):
        """Veritabanındaki tüm tamir kayıtlarında giris_tarihi alanı boşsa, tarih alanı ile doldurur"""
        try:
            self.db.cursor.execute("UPDATE tamirler SET giris_tarihi = tarih WHERE giris_tarihi IS NULL OR giris_tarihi = ''")
            self.db.conn.commit()
            messagebox.showinfo("Başarılı", "Eksik giris_tarihi alanları dolduruldu!")
        except Exception as e:
            messagebox.showerror("Hata", f"Eksik giris_tarihi doldurulamadı: {str(e)}")

    def setup_reports_tab(self):
        """Raporlar sekmesini profesyonel ve dinamik şekilde hazırlar"""
        self.reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.reports_frame, text="Raporlar")

        # --- Üstte filtre ve özet kutucukları ---
        self.report_top_frame = ttk.Frame(self.reports_frame)
        self.report_top_frame.pack(fill="x", padx=10, pady=5)

        # Rapor türü seçimi
        ttk.Label(self.report_top_frame, text="Rapor Türü:").grid(row=0, column=0, padx=5, pady=5)
        self.report_type = ttk.Combobox(self.report_top_frame, values=[
            "Tamir Raporu",
            "Müşteri Raporu",
            "Gelir Raporu",
            "Maliyet Analizi"
        ], state="readonly", width=20)
        self.report_type.grid(row=0, column=1, padx=5, pady=5)
        self.report_type.set("Tamir Raporu")
        self.report_type.bind("<<ComboboxSelected>>", lambda e: self.update_report_filters())

        # Tarih filtreleri
        ttk.Label(self.report_top_frame, text="Başlangıç:").grid(row=0, column=2, padx=5, pady=5)
        self.start_date = ttk.Entry(self.report_top_frame, width=12)
        self.start_date.grid(row=0, column=3, padx=5, pady=5)
        self.start_date.insert(0, datetime.now().strftime("%d.%m.%Y"))
        ttk.Label(self.report_top_frame, text="Bitiş:").grid(row=0, column=4, padx=5, pady=5)
        self.end_date = ttk.Entry(self.report_top_frame, width=12)
        self.end_date.grid(row=0, column=5, padx=5, pady=5)
        self.end_date.insert(0, datetime.now().strftime("%d.%m.%Y"))

        # Durum filtresi (sadece bazı raporlar için)
        self.status_label = ttk.Label(self.report_top_frame, text="Durum:")
        self.status_filter = ttk.Combobox(self.report_top_frame, values=[
            "Tümü", "Beklemede", "Parça Bekleniyor", "Tamir Ediliyor", "Test Ediliyor", "Tamamlandı", "İptal Edildi"
        ], state="readonly", width=15)
        self.status_filter.set("Tümü")

        # Rapor oluştur butonu
        ttk.Button(self.report_top_frame, text="Raporu Göster", command=self.generate_report).grid(row=0, column=10, padx=20, pady=5)

        # --- Özet kutucukları ---
        self.summary_frame = ttk.Frame(self.reports_frame)
        self.summary_frame.pack(fill="x", padx=10, pady=5)
        self.summary_labels = []  # Dinamik özet kutuları için

        # --- Tablo ve grafik ---
        self.report_middle_frame = ttk.Frame(self.reports_frame)
        self.report_middle_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tablo
        self.report_tree = ttk.Treeview(self.report_middle_frame, show="headings")
        self.report_tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(self.report_middle_frame, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="left", fill="y")

        # Grafik alanı
        self.report_graph_frame = ttk.Frame(self.report_middle_frame)
        self.report_graph_frame.pack(side="left", fill="both", expand=False, padx=10)
        self.report_graph_canvas = None

        # --- Dışa aktarım butonları ---
        export_frame = ttk.Frame(self.reports_frame)
        export_frame.pack(fill="x", padx=10, pady=5)
        ttk.Button(export_frame, text="Excel'e Aktar", command=self.export_report_excel).pack(side="left", padx=5)
        ttk.Button(export_frame, text="PDF'ye Aktar", command=self.export_report_pdf).pack(side="left", padx=5)

        # İlk açılışta filtreleri ve tabloyu ayarla
        self.update_report_filters()

    def update_report_filters(self):
        """Seçilen rapor türüne göre filtre ve tablo başlıklarını dinamik olarak ayarlar"""
        rapor_turu = self.report_type.get()
        # Durum filtresi sadece Tamir Raporu için gösterilsin
        if rapor_turu == "Tamir Raporu":
            self.status_label.grid(row=0, column=6, padx=5, pady=5)
            self.status_filter.grid(row=0, column=7, padx=5, pady=5)
        else:
            self.status_label.grid_remove()
            self.status_filter.grid_remove()
        # Tablo ve özet kutucuklarını temizle
        for lbl in getattr(self, 'summary_labels', []):
            lbl.destroy()
        self.summary_labels = []
        if hasattr(self, 'report_tree'):
            for col in self.report_tree["columns"]:
                self.report_tree.heading(col, text="")
            self.report_tree["columns"] = ()
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
        # Grafik alanını temizle
        if hasattr(self, 'report_graph_frame'):
            for widget in self.report_graph_frame.winfo_children():
                widget.destroy()
        self.report_graph_canvas = None

    def show_backup_manager(self):
        """Yedekleme yöneticisi penceresini gösterir"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Yedekleme Yöneticisi")
        dialog.geometry("600x400")
        dialog.transient(self.parent)
        dialog.grab_set()
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        # Yedek listesi
        tree = ttk.Treeview(
            main_frame,
            columns=("id", "tarih", "dosya_adi", "boyut", "aciklama"),
            show="headings"
        )
        tree.heading("id", text="ID")
        tree.heading("tarih", text="Tarih")
        tree.heading("dosya_adi", text="Dosya Adı")
        tree.heading("boyut", text="Boyut")
        tree.heading("aciklama", text="Açıklama")
        tree.column("id", width=0, stretch=False)
        tree.column("tarih", width=150)
        tree.column("dosya_adi", width=220)
        tree.column("boyut", width=100)
        tree.column("aciklama", width=300)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        # --- Yedekleri buraya ekle ---
        self.refresh_backup_manager_tree(tree)
        tree.tag_configure("eksik", background="tomato")
        # Butonlar
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        def yedek_al():
            try:
                self.db.yedek_al("Manuel yedekleme")
                self.refresh_backups()
                self.refresh_backup_manager_tree(tree)
                yedekler = self.db.tum_yedeklemeler()
                if yedekler:
                    dosya_adi = yedekler[0]["dosya_adi"]
                    dosya_yolu = os.path.join("backups", dosya_adi)
                    result = self.send_backup_email(dosya_yolu)
                    if result:
                        messagebox.showinfo("Başarılı", "Yedek alındı ve e-posta gönderildi")
                    else:
                        messagebox.showwarning("Uyarı", "Yedek alındı fakat e-posta gönderilemedi!")
                else:
                    messagebox.showinfo("Başarılı", "Yedek alındı")
            except Exception as e:
                messagebox.showerror("Hata", str(e))
        def yedek_yukle():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "Lütfen bir yedek seçin")
                return
            if messagebox.askyesno("Onay", "Seçili yedeği geri yüklemek istediğinize emin misiniz?"):
                try:
                    item_data = tree.item(selected[0])
                    if not item_data or "values" not in item_data or not item_data["values"]:
                        messagebox.showerror("Hata", "Seçili yedek satırında veri yok!")
                        return
                    yedek_id = item_data["values"][0]
                    dosya_adi = item_data["values"][2]
                    dosya_yolu = os.path.join("backups", dosya_adi)
                    if not os.path.exists(dosya_yolu):
                        messagebox.showerror("Hata", f"Yedek dosyası fiziksel olarak bulunamadı:\n{dosya_yolu}")
                        return
                    self.db.yedek_geri_yukle(yedek_id)
                    self.refresh_backups()
                    self.refresh_backup_manager_tree(tree)
                    messagebox.showinfo("Başarılı", "Yedek geri yüklendi")
                except Exception as e:
                    messagebox.showerror("Hata", str(e))
        def yedek_sil():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "Lütfen bir yedek seçin")
                return
            if messagebox.askyesno("Onay", "Seçili yedeği silmek istediğinize emin misiniz?"):
                try:
                    item_data = tree.item(selected[0])
                    if not item_data or "values" not in item_data or not item_data["values"]:
                        messagebox.showerror("Hata", "Seçili yedek satırında veri yok!")
                        return
                    yedek_id = item_data["values"][0]
                    self.db.yedek_sil(yedek_id)
                    self.refresh_backups()
                    self.refresh_backup_manager_tree(tree)
                    messagebox.showinfo("Başarılı", "Yedek silindi")
                except Exception as e:
                    messagebox.showerror("Hata", str(e))
        def yedek_eposta():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "Lütfen bir yedek seçin")
                return
            item_data = tree.item(selected[0])
            if not item_data or "values" not in item_data or not item_data["values"]:
                messagebox.showerror("Hata", "Seçili yedek satırında veri yok!")
                return
            dosya_adi = item_data["values"][2]
            dosya_yolu = os.path.join("backups", dosya_adi)
            result = self.send_backup_email(dosya_yolu)
            if result:
                messagebox.showinfo("Başarılı", "Yedek e-posta ile gönderildi")
            else:
                messagebox.showwarning("Uyarı", "E-posta gönderilemedi!")
        ttk.Button(button_frame, text="Yedek Al", command=yedek_al).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Geri Yükle", command=yedek_yukle).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Sil", command=yedek_sil).pack(side="left", padx=5)
        ttk.Button(button_frame, text="E-posta ile Gönder", command=yedek_eposta).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Kapat", command=dialog.destroy).pack(side="left", padx=5)
        # Yedek listesini güncelle
        self.refresh_backup_manager_tree(tree)

    def refresh_backup_manager_tree(self, tree):
        # Yedekleme yöneticisi ekranındaki tabloyu günceller
        for item in tree.get_children():
            tree.delete(item)
        yedekler = self.db.tum_yedeklemeler()
        for yedek in yedekler:
            try:
                tarih = datetime.strptime(yedek["tarih"], "%Y-%m-%d %H:%M:%S") + timedelta(hours=3)
                tarih_str = tarih.strftime("%d.%m.%Y %H:%M:%S")
            except Exception:
                tarih_str = yedek["tarih"]
            boyut_kb = int(yedek["boyut"]) // 1024 if yedek["boyut"] else 0
            dosya_adi = yedek["dosya_adi"]
            dosya_yolu = os.path.join("backups", dosya_adi)
            exists = os.path.exists(dosya_yolu)
            row_id = tree.insert(
                "", tk.END,
                values=(yedek["id"], tarih_str, dosya_adi, f"{boyut_kb} KB", yedek["aciklama"])
            )
            if not exists:
                tree.item(row_id, tags=("eksik",))
        tree.tag_configure("eksik", background="tomato")

    def refresh_backup_manager(self):
        # Yedekleme yöneticisi ekranındaki tabloyu güncellemek için kullanılacak (varsa)
        pass

    def giris_tarihi_ileri_al(self):
        """Tüm tamir kayıtlarının giris_tarihi alanını 3 saat ileri alır"""
        try:
            self.db.cursor.execute("UPDATE tamirler SET giris_tarihi = datetime(giris_tarihi, '+3 hours') WHERE giris_tarihi IS NOT NULL AND giris_tarihi != ''")
            self.db.conn.commit()
            messagebox.showinfo("Başarılı", "Tüm giris_tarihi alanları 3 saat ileri alındı!")
        except Exception as e:
            messagebox.showerror("Hata", f"Saat güncelleme başarısız: {str(e)}")

    def setup_about_tab(self):
        """Hakkında sekmesini ekler (profesyonel görünüm)"""
        self.about_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.about_frame, text="Hakkında")
        # Başlık
        title = ttk.Label(self.about_frame, text="Tamir Atölyesi Yönetim Sistemi", font=("Arial", 18, "bold"), foreground="#2a4d69")
        title.pack(padx=20, pady=(30,10), anchor="n")
        # Alt başlık
        subtitle = ttk.Label(self.about_frame, text="Sürüm 1.0", font=("Arial", 12, "italic"), foreground="#4b86b4")
        subtitle.pack(padx=20, pady=(0,10), anchor="n")
        # Logo veya ikon (varsa)
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "../icon.ico")
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                img = img.resize((64, 64))
                logo = ImageTk.PhotoImage(img)
                logo_label = tk.Label(self.about_frame, image=logo)
                logo_label.pack(pady=(0, 10))
                setattr(logo_label, "image", logo)
        except Exception:
            pass
        # Açıklama kutusu
        desc_frame = ttk.LabelFrame(self.about_frame, text="Program Hakkında", padding="15")
        desc_frame.pack(padx=30, pady=10, fill="x", anchor="n")
        desc = ("Bu yazılım, tamir atölyeleri için müşteri ve tamir yönetimini kolaylaştırmak amacıyla geliştirilmiştir.\n"
                "Kullanıcı dostu arayüzü ve kapsamlı raporlama özellikleriyle işlerinizi kolaylaştırır.\n\n"
                "© 2025 Tüm hakları saklıdır.\nGeliştirici: Mehmet Said ÖZSOY\nİletişim: 0 541 781 33 68")
        ttk.Label(desc_frame, text=desc, font=("Arial", 11), justify="left", wraplength=500).pack(anchor="w")

    def setup_manual_tab(self):
        """Kullanım Bilgisi (Manuel) sekmesini ekler (profesyonel görünüm)"""
        self.manual_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.manual_frame, text="Kullanım Bilgisi")
        frame = ttk.Frame(self.manual_frame, padding="10")
        frame.pack(fill="both", expand=True)
        # Başlık
        title = ttk.Label(frame, text="Kullanıcı Kılavuzu", font=("Arial", 16, "bold"), foreground="#2a4d69")
        title.pack(pady=(10, 10), anchor="n")
        # Açıklama
        ttk.Label(frame, text="Sistemin temel kullanım adımları aşağıda özetlenmiştir:", font=("Arial", 11, "italic"), foreground="#4b86b4").pack(anchor="w", padx=10)
        # Madde kutuları
        steps = [
            "Yeni müşteri eklemek için: 'Yeni Müşteri' butonunu veya menüsünü kullanın.",
            "Tamir kaydı eklemek için: 'Yeni Tamir' butonunu veya menüsünü kullanın.",
            "Yedekleme işlemleri için: Ayarlar sekmesinden veya 'Yedekleme Yöneticisi'nden işlem yapabilirsiniz.",
            "Raporlar sekmesinden çeşitli raporları oluşturabilirsiniz.",
            "Detaylı bilgi için geliştiriciye başvurun."
        ]
        for step in steps:
            box = ttk.Frame(frame)
            box.pack(fill="x", padx=20, pady=5, anchor="w")
            icon = ttk.Label(box, text="•", font=("Arial", 14, "bold"), foreground="#63ace5")
            icon.pack(side="left", padx=(0,8))
            ttk.Label(box, text=step, font=("Arial", 11), wraplength=600, justify="left").pack(side="left", fill="x", expand=True)
        # Ek açıklama
        ttk.Label(frame, text="\nBu alanı dilediğiniz gibi genişletebilir, yeni başlıklar ve açıklamalar ekleyebilirsiniz.", font=("Arial", 10), foreground="#888").pack(anchor="w", padx=10, pady=(10,0))

    def setup_admin_panel_tab(self):
        """Yönetim Paneli sekmesini ekler (sadece admin)"""
        self.admin_panel_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.admin_panel_frame, text="Yönetim Paneli")
        # Başlık
        ttk.Label(self.admin_panel_frame, text="Kullanıcı Yönetimi", font=("Arial", 16, "bold"), foreground="#2a4d69").pack(pady=(10, 10))
        # Kullanıcı listesi
        user_frame = ttk.LabelFrame(self.admin_panel_frame, text="Kullanıcılar", padding="10")
        user_frame.pack(fill="x", padx=20, pady=10)
        self.user_tree = ttk.Treeview(user_frame, columns=("kullanici_adi", "rol"), show="headings", height=6)
        self.user_tree.heading("kullanici_adi", text="Kullanıcı Adı")
        self.user_tree.heading("rol", text="Rol")
        self.user_tree.column("kullanici_adi", width=150)
        self.user_tree.column("rol", width=80)
        self.user_tree.pack(side="left", fill="x", expand=True)
        scrollbar = ttk.Scrollbar(user_frame, orient="vertical", command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        # Butonlar
        btn_frame = ttk.Frame(self.admin_panel_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Kullanıcı Ekle", command=self.add_user_dialog).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Kullanıcı Sil", command=self.delete_user).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Şifre Değiştir", command=self.change_user_password_dialog).pack(side="left", padx=5)
        # Listeyi doldur
        self.refresh_user_list()

    def refresh_user_list(self):
        """Kullanıcı listesini günceller"""
        for item in self.user_tree.get_children():
            self.user_tree.delete(item)
        users = self.db.tum_kullanicilar()
        for u in users:
            rol = "Admin" if u[2] == 0 else "Personel"
            self.user_tree.insert("", "end", values=(u[1], rol))

    def add_user_dialog(self):
        """Yeni kullanıcı ekleme penceresi"""
        dialog = tk.Toplevel(self.parent)
        dialog.title("Kullanıcı Ekle")
        dialog.geometry("300x220")
        dialog.transient(self.parent)
        dialog.grab_set()
        ttk.Label(dialog, text="Kullanıcı Adı:").pack(pady=5)
        username = ttk.Entry(dialog)
        username.pack(pady=5)
        ttk.Label(dialog, text="Şifre:").pack(pady=5)
        password = ttk.Entry(dialog, show="*")
        password.pack(pady=5)
        ttk.Label(dialog, text="Rol:").pack(pady=5)
        role = ttk.Combobox(dialog, values=["Admin", "Personel"], state="readonly")
        role.set("Personel")
        role.pack(pady=5)
        def ekle():
            ad = username.get().strip()
            sif = password.get().strip()
            rol = 0 if role.get() == "Admin" else 1
            if not ad or not sif:
                messagebox.showwarning("Uyarı", "Kullanıcı adı ve şifre zorunludur!", parent=dialog)
                return
            sonuc = self.db.kullanici_ekle(ad, sif, rol)
            if sonuc:
                messagebox.showinfo("Başarılı", "Kullanıcı eklendi!", parent=dialog)
                self.refresh_user_list()
                dialog.destroy()
            else:
                messagebox.showerror("Hata", "Kullanıcı eklenemedi! (Aynı kullanıcı adı olabilir)", parent=dialog)
        ttk.Button(dialog, text="Ekle", command=ekle).pack(pady=10)
        ttk.Button(dialog, text="İptal", command=dialog.destroy).pack()

    def delete_user(self):
        """Seçili kullanıcıyı siler"""
        selected = self.user_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen silinecek kullanıcıyı seçin!")
            return
        username = self.user_tree.item(selected[0])["values"][0]
        if username == "admin":
            messagebox.showwarning("Uyarı", "Admin kullanıcısı silinemez!")
            return
        if messagebox.askyesno("Onay", f"{username} adlı kullanıcıyı silmek istediğinize emin misiniz?"):
            sonuc = self.db.kullanici_sil(username)
            if sonuc:
                messagebox.showinfo("Başarılı", "Kullanıcı silindi!")
                self.refresh_user_list()
            else:
                messagebox.showerror("Hata", "Kullanıcı silinemedi!")

    def change_user_password_dialog(self):
        """Seçili kullanıcının şifresini değiştir"""
        selected = self.user_tree.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Lütfen şifresi değiştirilecek kullanıcıyı seçin!")
            return
        username = self.user_tree.item(selected[0])["values"][0]
        dialog = tk.Toplevel(self.parent)
        dialog.title("Şifre Değiştir")
        dialog.geometry("300x180")
        dialog.transient(self.parent)
        dialog.grab_set()
        ttk.Label(dialog, text=f"Kullanıcı: {username}").pack(pady=10)
        ttk.Label(dialog, text="Yeni Şifre:").pack(pady=5)
        new_pass = ttk.Entry(dialog, show="*")
        new_pass.pack(pady=5)
        def degistir():
            sif = new_pass.get().strip()
            if not sif:
                messagebox.showwarning("Uyarı", "Yeni şifre zorunludur!", parent=dialog)
                return
            sonuc = self.db.kullanici_sifre_guncelle(username, sif)
            if sonuc:
                messagebox.showinfo("Başarılı", "Şifre değiştirildi!", parent=dialog)
                dialog.destroy()
            else:
                messagebox.showerror("Hata", "Şifre değiştirilemedi!", parent=dialog)
        ttk.Button(dialog, text="Değiştir", command=degistir).pack(pady=10)
        ttk.Button(dialog, text="İptal", command=dialog.destroy).pack()
